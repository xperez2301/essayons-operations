import os
import csv
import json
import math
import re
import shutil
import requests
import zipfile
import subprocess
import sys
import time
from html import escape
from io import BytesIO
from functools import wraps
import secrets
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from flask import Flask, jsonify, render_template, request, send_file, abort, redirect, url_for, session, g, has_request_context
from openpyxl import load_workbook
from pypdf import PdfReader
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

# Playwright is only needed for the RMS scraping features. It is heavy and not
# always available on a fresh App Service worker, so we import it lazily and let
# the rest of the app run even if it is missing.
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_AVAILABLE = True
except Exception:  # pragma: no cover
    PLAYWRIGHT_AVAILABLE = False

    class PlaywrightTimeoutError(Exception):
        pass

    def sync_playwright(*args, **kwargs):
        raise RuntimeError(
            "Playwright is not installed on this server, so RMS automation is "
            "disabled. Install it with: pip install playwright && playwright install chromium"
        )

app = Flask(__name__)

# Load a local .env file for development (no-op if python-dotenv or the file is
# absent). On Azure, real values come from Application settings instead.
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# ---------------------------------------------------------------------------
# Configuration. Everything sensitive comes from environment variables first
# (set these in Azure -> Configuration -> Application settings) and only falls
# back to a default for local development.
# ---------------------------------------------------------------------------
app.secret_key = os.environ.get("SECRET_KEY")
if not app.secret_key:
    raise RuntimeError(
        "SECRET_KEY is required. Set it to a long random value in the hosting "
        "environment before starting EOMS."
    )

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("SESSION_COOKIE_SECURE", "1") == "1",
    SEND_FILE_MAX_AGE_DEFAULT=3600,
)

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")

BASE_DIR = Path(__file__).resolve().parent

# DATA_DIR / BOL_DIR can be pointed at a persistent location so a redeploy does
# not wipe live data. On Azure App Service (Linux) use /home/data, which
# survives restarts and deployments. Locally it defaults to ./data.
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR / "data")))
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", str(BASE_DIR / "uploads")))
BOL_DIR = Path(os.environ.get("BOL_DIR", str(BASE_DIR / "bol_files")))

STORES_FILE = DATA_DIR / "stores.json"
ROUTES_FILE = DATA_DIR / "routes.json"
AUDIT_FILE = DATA_DIR / "audit_log.json"
SETTINGS_FILE = DATA_DIR / "settings.json"
SYNC_HISTORY_FILE = DATA_DIR / "sync_history.json"
RMS_QUEUE_FILE = DATA_DIR / "rms_queue.json"
USERS_FILE = DATA_DIR / "users.json"

MAX_PAYLOAD = 25001
WARNING_PAYLOAD = 22000
PIECES_PER_RACK = 19
RATE_PER_PIECE = 0.95
DRIVER_PAY_PER_PIECE = 0.30
DEFAULT_RACK_WEIGHT = 200

HUBS = {
    "San Antonio": {
        "lat": 29.5353,
        "lng": -98.4188,
        "address": "10536 Sentinel Dr, San Antonio, TX 78217"
    },
    "Houston": {
        "lat": 30.0147,
        "lng": -95.4306,
        "address": "403 Century Plaza Dr, Houston, TX 77073"
    },
    "Dallas": {
        "lat": 32.6386,
        "lng": -96.8662,
        "address": "2777 Danieldale Rd, Dallas, TX 75237"
    }
}

CITY_COORDS = {
    "San Antonio": (29.4241, -98.4936), "Houston": (29.7604, -95.3698),
    "Dallas": (32.7767, -96.7970), "Irving": (32.8140, -96.9489),
    "Killeen": (31.1171, -97.7278), "Austin": (30.2672, -97.7431),
    "Waco": (31.5493, -97.1467), "San Marcos": (29.8833, -97.9414),
    "New Braunfels": (29.7030, -98.1245), "Selma": (29.5844, -98.3058),
    "Kyle": (29.9891, -97.8772), "Corpus Christi": (27.8006, -97.3964),
    "Brownsville": (25.9017, -97.4975), "McAllen": (26.2034, -98.2300),
    "Beaumont": (30.0802, -94.1266), "Marble Falls": (30.5782, -98.2728),
    "Abilene": (32.4487, -99.7331), "Fort Worth": (32.7555, -97.3308),
    "Lubbock": (33.5779, -101.8552)
}

DIRS_READY = False

def ensure_dirs():
    global DIRS_READY
    if DIRS_READY:
        return
    DATA_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(exist_ok=True)
    for root in ["Imported", "Assigned", "Dispatched", "Completed", "Need_Review", "RMS_Backup", "RMS_Closed"]:
        (BOL_DIR / root).mkdir(parents=True, exist_ok=True)
    if not USERS_FILE.exists() and not ADMIN_PASSWORD:
        raise RuntimeError(
            "ADMIN_PASSWORD is required on first startup so EOMS can create "
            "the initial administrator account."
        )
    for path, default in [
        (STORES_FILE, []), (ROUTES_FILE, []), (AUDIT_FILE, []),
        (SETTINGS_FILE, {"rms_username": "", "rms_password": "", "rms_password_saved": False, "remember_rms_credentials": True, "last_rms_login": "", "rms_connection_status": "Not Tested", "rms_login_url": "https://rms.reusability.com/login", "rms_bol_url": "https://rms.reusability.com/bills-of-lading", "azure_maps_key": "", "map_default_type": "satellite", "map_default_zoom": 7, "map_board_default_open": False, "map_live_refresh_seconds": 30, "due_red_days": 4, "due_amber_days": 7}),
        (SYNC_HISTORY_FILE, []),
        (RMS_QUEUE_FILE, []),
        (USERS_FILE, {"users":[{"id":"admin","username":ADMIN_USERNAME,"password":hash_password(ADMIN_PASSWORD),"role":"Admin","assigned_cities":["All"],"active":True,"created_at":"system"}]})
    ]:
        if not path.exists():
            path.write_text(json.dumps(default, indent=2), encoding="utf-8")
    DIRS_READY = True

def read_json(path):
    ensure_dirs()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        data = [] if path != SETTINGS_FILE else {}
    # Environment variables win over anything stored on disk for secrets, so the
    # repository never needs to contain real credentials.
    if path == SETTINGS_FILE and isinstance(data, dict):
        defaults = {
            "map_default_type": "satellite",
            "map_default_zoom": 7,
            "map_board_default_open": False,
            "map_live_refresh_seconds": 30,
            "due_red_days": 4,
            "due_amber_days": 7,
        }
        for key, value in defaults.items():
            data.setdefault(key, value)
        env_overlay = {
            "rms_username": os.environ.get("RMS_USERNAME"),
            "rms_password": os.environ.get("RMS_PASSWORD"),
            "azure_maps_key": os.environ.get("AZURE_MAPS_KEY"),
        }
        for key, value in env_overlay.items():
            if value:
                data[key] = value
        if data.get("rms_password"):
            data["rms_password_saved"] = True
    return data

def write_json(path, data):
    ensure_dirs()
    # Atomic write: write to a temp file then replace, so a crash mid-write can
    # never leave a half-written (corrupt) JSON file behind.
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2), encoding="utf-8")
    os.replace(tmp, path)


# ---------------------------------------------------------------------------
# Password handling. Stored passwords are salted hashes (werkzeug). Any legacy
# plaintext password found on disk is verified once and then transparently
# upgraded to a hash, so existing logins keep working after the upgrade.
# ---------------------------------------------------------------------------
def hash_password(raw):
    return generate_password_hash(raw or "")

def looks_hashed(value):
    return isinstance(value, str) and value.startswith(("pbkdf2:", "scrypt:", "argon2"))

def verify_password(stored, provided):
    """Return (is_valid, needs_upgrade)."""
    if stored is None:
        return False, False
    if looks_hashed(stored):
        return check_password_hash(stored, provided or ""), False
    # Legacy plaintext comparison (constant work) — valid once, then upgrade.
    return secrets.compare_digest(str(stored), str(provided or "")), True

def migrate_user_passwords():
    """One-time, idempotent: hash any plaintext passwords stored in users.json."""
    try:
        data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return
    changed = False
    for user in data.get("users", []):
        pw = user.get("password")
        if pw and not looks_hashed(pw):
            user["password"] = hash_password(pw)
            changed = True
    if changed:
        write_json(USERS_FILE, data)
        print("[EOMS] Upgraded stored user passwords to hashed form.")

def audit(action, details):
    log = read_json(AUDIT_FILE)
    log.append({"id": str(uuid4()), "time": datetime.now().isoformat(timespec="seconds"), "action": action, "details": details})
    write_json(AUDIT_FILE, log)

def local_rms_import_authorized():
    token = clean(os.environ.get("LOCAL_RMS_IMPORT_TOKEN"))
    if not token:
        return False
    auth = clean(request.headers.get("Authorization"))
    supplied = ""
    if auth.lower().startswith("bearer "):
        supplied = clean(auth[7:])
    supplied = supplied or clean(request.headers.get("X-EOMS-Import-Token"))
    return bool(supplied) and secrets.compare_digest(supplied, token)

def clean(value):
    return "" if value is None else str(value).strip()

def num(value, default=0):
    try:
        if value in [None, ""]:
            return default
        return float(str(value).replace(",", "").strip())
    except Exception:
        return default

def safe_part(value):
    value = re.sub(r"[^A-Za-z0-9_-]+", "", clean(value).replace(" ", ""))
    return value[:35] or "Unknown"

def month_folder(root, when=None):
    when = when or datetime.now()
    folder = BOL_DIR / root / str(when.year) / f"{when.month:02d}_{when.strftime('%B')}"
    folder.mkdir(parents=True, exist_ok=True)
    return folder

def move_pdf(current_path, destination_root):
    current = Path(current_path)
    if not current.exists():
        return str(current_path)
    dest = month_folder(destination_root) / current.name
    if dest.exists():
        dest = dest.with_name(dest.stem + "_" + datetime.now().strftime("%H%M%S") + dest.suffix)
    shutil.move(str(current), str(dest))
    return str(dest)

def path_inside(child, parent):
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False

def delete_saved_bol_files(item):
    deleted = []
    failed = []
    candidates = []
    for key in ("pdf_path", "printable_path"):
        path = clean(item.get(key))
        if path:
            candidates.append(Path(path))

    bol = safe_part(item.get("bol"))
    if bol and bol != "Unknown":
        candidates.extend(BOL_DIR.rglob(f"*{bol}*.pdf"))
        candidates.extend(BOL_DIR.rglob(f"*{bol}*.html"))

    seen = set()
    for path in candidates:
        try:
            resolved = path.resolve()
        except Exception:
            continue
        if str(resolved) in seen:
            continue
        seen.add(str(resolved))
        if not (path_inside(resolved, BOL_DIR) or path_inside(resolved, UPLOAD_DIR)):
            continue
        if resolved.exists() and resolved.is_file():
            try:
                resolved.unlink()
                deleted.append(str(resolved))
            except Exception as exc:
                failed.append({"path": str(resolved), "error": str(exc)[:160]})
    return {"deleted": deleted, "failed": failed}

def bol_edit_summary_html(record):
    fields = [
        ("BOL", record.get("bol")),
        ("Status", record.get("status") or record.get("queue_status")),
        ("Store", record.get("store_name") or record.get("origin_name")),
        ("Origin", record.get("origin")),
        ("Address", record.get("address") or record.get("origin_address")),
        ("City/State/ZIP", " ".join(x for x in [
            clean(record.get("city") or record.get("origin_city")),
            clean(record.get("state") or record.get("origin_state")),
            clean(record.get("zip") or record.get("origin_zip"))
        ] if x)),
        ("Contact", record.get("contact") or record.get("origin_contact")),
        ("Hub", record.get("hub")),
        ("Expected Racks", record.get("expected_racks")),
        ("Weight", record.get("weight")),
        ("Due Date", record.get("due_date")),
        ("Assigned Date", record.get("assigned_date")),
        ("Updated", record.get("updated_at")),
    ]
    rows = []
    for label, value in fields:
        value = clean(value)
        if value:
            rows.append(f"<tr><th>{escape(label)}</th><td>{escape(value)}</td></tr>")
    if not rows:
        return ""
    return f"""
<section class="eoms-bol-edits">
  <h2>EOMS Updated BOL Details</h2>
  <table>{''.join(rows)}</table>
</section>
"""

def bol_print_styles():
    return """
<style>
.eoms-bol-toolbar {
  position: sticky;
  top: 0;
  background: #0f172a;
  color: white;
  padding: 10px;
  display: flex;
  align-items: center;
  gap: 10px;
  z-index: 9999;
  font-family: Arial, sans-serif;
}
.eoms-bol-toolbar button {
  padding: 8px 12px;
  font-weight: 800;
  cursor: pointer;
}
.eoms-bol-edits {
  margin: 0 0 12px;
  padding: 10px 12px;
  border: 2px solid #0f172a;
  font-family: Arial, sans-serif;
  page-break-inside: avoid;
}
.eoms-bol-edits h2 {
  margin: 0 0 8px;
  font-size: 16px;
  letter-spacing: 0;
}
.eoms-bol-edits table {
  width: 100%;
  border-collapse: collapse;
  font-size: 12px;
}
.eoms-bol-edits th,
.eoms-bol-edits td {
  border: 1px solid #cbd5e1;
  padding: 5px 6px;
  text-align: left;
  vertical-align: top;
}
.eoms-bol-edits th {
  width: 145px;
  background: #f1f5f9;
}
.eoms-bol-frame {
  width: 100%;
  height: calc(100vh - 190px);
  border: 0;
}
@media print {
  .eoms-bol-toolbar { display: none !important; }
  body { margin: 0.25in; }
  .eoms-bol-frame { height: 9.5in; }
}
</style>
"""

def render_saved_bol(record, path, auto_print=False):
    if request.args.get("raw") == "1":
        return send_file(Path(path), as_attachment=False)

    summary = bol_edit_summary_html(record)
    bol_id = record.get("id") or clean(record.get("bol"))
    raw_url = url_for("bol_view", store_id=bol_id, raw=1)
    title = f"BOL {escape(clean(record.get('bol')))}"
    print_script = "<script>window.addEventListener('load', function(){ setTimeout(function(){ window.print(); }, 700); });</script>" if auto_print else ""
    toolbar = f"""
<div class="eoms-bol-toolbar">
  <button onclick="window.print()">Print to Office Printer</button>
  <button onclick="window.close()">Close</button>
  <span>{title}</span>
</div>
"""

    if str(path).lower().endswith(".html"):
        content = Path(path).read_text(encoding="utf-8", errors="ignore")
        return f"""<!doctype html>
<html>
<head>
<title>Print {title}</title>
{bol_print_styles()}
</head>
<body>
{toolbar if auto_print else ""}
{summary}
{content}
{print_script}
</body>
</html>"""

    return f"""<!doctype html>
<html>
<head>
<title>Print {title}</title>
{bol_print_styles()}
</head>
<body>
{toolbar}
{summary}
<iframe class="eoms-bol-frame" src="{escape(raw_url)}"></iframe>
{print_script}
</body>
</html>"""

def bol_duplicate_key(item):
    bol = clean(item.get("bol"))
    origin = clean(item.get("origin"))
    return bol, origin

def is_duplicate_bol(item, existing_keys, existing_bols):
    bol, origin = bol_duplicate_key(item)
    if bol and origin and (bol, origin) in existing_keys:
        return True
    # Also protect against the same BOL coming back with a missing/changed origin.
    if bol and bol in existing_bols:
        return True
    return False

def track_bol_key(item, existing_keys, existing_bols):
    bol, origin = bol_duplicate_key(item)
    if bol or origin:
        existing_keys.add((bol, origin))
    if bol:
        existing_bols.add(bol)

def completion_summary_for_stores(stores):
    expected = round(sum(num(s.get("expected_racks")) for s in stores), 2)
    collected = round(sum(num(s.get("collected_racks")) for s in stores), 2)
    expected_pieces = round(sum(num(s.get("expected_pieces")) or (num(s.get("expected_racks")) * PIECES_PER_RACK) for s in stores), 2)
    collected_pieces = round(sum(num(s.get("collected_pieces")) for s in stores), 2)
    return {
        "stores": len(stores),
        "expected_racks": expected,
        "collected_racks": collected,
        "variance": round(collected - expected, 2),
        "expected_pieces": expected_pieces,
        "collected_pieces": collected_pieces,
        "pieces_variance": round(collected_pieces - expected_pieces, 2),
        "completed_at": datetime.now().isoformat(timespec="seconds"),
    }

def path_health(path):
    info = {"path": str(path), "exists": path.exists(), "writable": False, "error": ""}
    try:
        path.mkdir(parents=True, exist_ok=True)
        test_file = path / f".health_{uuid4().hex}"
        test_file.write_text("ok", encoding="utf-8")
        test_file.unlink(missing_ok=True)
        info["exists"] = True
        info["writable"] = True
    except Exception as exc:
        info["error"] = str(exc)[:180]
    return info

def run_playwright_install_command(args, timeout=300):
    result = subprocess.run(
        [sys.executable, "-m", "playwright", *args],
        env=os.environ.copy(),
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Playwright command failed: python -m playwright {' '.join(args)}. "
            f"stdout: {(result.stdout or '')[-800:]} stderr: {(result.stderr or '')[-1200:]}"
        )
    return True

def install_playwright_chromium():
    """Install Chromium into persistent Azure storage when Playwright has no browser binary yet."""
    browser_path = os.environ.get("PLAYWRIGHT_BROWSERS_PATH") or "/home/playwright"
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = browser_path
    Path(browser_path).mkdir(parents=True, exist_ok=True)
    return run_playwright_install_command(["install", "chromium"])

def install_playwright_system_deps():
    """Install Linux shared libraries needed by Chromium on fresh App Service workers."""
    return run_playwright_install_command(["install-deps", "chromium"], timeout=600)

def is_missing_browser_binary(message):
    return "Executable doesn't exist" in message or "please run playwright install" in message

def is_missing_browser_deps(message):
    return (
        "Host system is missing dependencies" in message or
        "playwright install-deps" in message or
        "libglib2.0-0" in message or
        "libnss3" in message
    )

def launch_chromium_with_repair(playwright, headless=True):
    try:
        return playwright.chromium.launch(headless=headless)
    except Exception as exc:
        message = str(exc)
        if not is_missing_browser_binary(message) and not is_missing_browser_deps(message):
            raise
        if is_missing_browser_deps(message):
            install_playwright_system_deps()
        install_playwright_chromium()
        return playwright.chromium.launch(headless=headless)

def rms_browser_choice():
    return clean(os.environ.get("RMS_BROWSER") or "chromium").lower()

def rms_manual_login_enabled():
    return clean(os.environ.get("RMS_MANUAL_LOGIN")).lower() in {"1", "true", "yes", "on"}

def rms_edge_profile_dir():
    configured = clean(os.environ.get("RMS_EDGE_PROFILE_DIR")) or clean(os.environ.get("RMS_PROFILE_DIR"))
    if configured:
        return Path(configured)
    profile_dir = BASE_DIR / "runtime" / "rms_edge_profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    return profile_dir

def rms_chromium_profile_dir():
    configured = clean(os.environ.get("RMS_CHROMIUM_PROFILE_DIR"))
    if configured:
        return Path(configured)
    profile_dir = BASE_DIR / "runtime" / "rms_chromium_profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    return profile_dir

def rms_cdp_endpoint():
    return clean(os.environ.get("RMS_CDP_ENDPOINT") or "http://127.0.0.1:9222")

def launch_rms_browser_context(playwright, headless=True):
    """Launch the browser context used for RMS.

    RMS_BROWSER=edge uses installed Microsoft Edge with a persistent EOMS-owned
    profile so the RMS session/cookies can be reused between imports. This is
    safer than attaching to the user's active Default profile, which Edge may
    lock while it is open.
    """
    choice = rms_browser_choice()
    common_kwargs = {
        "user_agent": RMS_BROWSER_USER_AGENT,
        "viewport": {"width": 1366, "height": 768},
        "locale": "en-US",
        "timezone_id": "America/Chicago",
        "ignore_https_errors": True,
        "extra_http_headers": {
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        },
    }
    if choice in {"edge", "msedge", "microsoft-edge"}:
        try:
            context = playwright.chromium.launch_persistent_context(
                user_data_dir=str(rms_edge_profile_dir()),
                channel="msedge",
                headless=headless,
                args=["--start-maximized"],
                **common_kwargs,
            )
            return None, context
        except Exception as exc:
            edge_error = str(exc)[:500]
            try:
                context = playwright.chromium.launch_persistent_context(
                    user_data_dir=str(rms_chromium_profile_dir()),
                    headless=headless,
                    args=["--start-maximized"],
                    **common_kwargs,
                )
                return None, context
            except Exception as fallback_exc:
                raise RuntimeError(
                    "EOMS could not launch Microsoft Edge for RMS, and the Chromium fallback also failed. "
                    "Close all RMS/Edge automation windows and try again. Edge details: "
                    + edge_error + " Chromium details: " + str(fallback_exc)[:500]
                )

    if choice in {"cdp", "edge-cdp", "remote-edge"}:
        try:
            browser = playwright.chromium.connect_over_cdp(rms_cdp_endpoint())
            contexts = browser.contexts
            context = contexts[0] if contexts else browser.new_context(**common_kwargs)
            return browser, context
        except Exception as exc:
            raise RuntimeError(
                "EOMS could not connect to the user-launched RMS Edge browser. "
                "Run START_RMS_EDGE_DEBUG.bat first, log into RMS in that window, then run Auto Grab. "
                "Details: " + str(exc)[:500]
            )

    browser = launch_chromium_with_repair(playwright, headless=headless)
    return browser, new_rms_context(browser)

def close_rms_browser(browser, context):
    if rms_manual_login_enabled():
        return
    try:
        if context:
            context.close()
    except Exception:
        pass
    try:
        if browser:
            browser.close()
    except Exception:
        pass

def rms_headless(default=True):
    value = clean(os.environ.get("RMS_HEADLESS"))
    if value == "":
        return default
    return value.lower() not in {"0", "false", "no", "off"}

def rms_debug_hold(page):
    try:
        seconds = int(os.environ.get("RMS_DEBUG_HOLD_SECONDS", "0") or 0)
    except Exception:
        seconds = 0
    if seconds > 0:
        page.wait_for_timeout(seconds * 1000)

def miles_between(lat1, lng1, lat2, lng2):
    radius = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2-lat1), math.radians(lng2-lng1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return radius * (2 * math.atan2(math.sqrt(a), math.sqrt(1-a)))


def route_miles(points):
    if len(points) < 2:
        return 0
    total = 0
    for i in range(len(points) - 1):
        total += miles_between(points[i]["lat"], points[i]["lng"], points[i+1]["lat"], points[i+1]["lng"])
    return round(total, 1)

def order_nearest_from_hub(hub_name, selected_stores):
    hub = HUBS.get(hub_name) or HUBS["San Antonio"]
    current = {"lat": hub["lat"], "lng": hub["lng"]}
    remaining = selected_stores[:]
    ordered = []
    while remaining:
        next_stop = min(remaining, key=lambda s: miles_between(current["lat"], current["lng"], num(s.get("lat")), num(s.get("lng"))))
        ordered.append(next_stop)
        remaining.remove(next_stop)
        current = {"lat": num(next_stop.get("lat")), "lng": num(next_stop.get("lng"))}
    return ordered

def calculate_route_metrics(ordered_stores, hub_name):
    hub = HUBS.get(hub_name) or HUBS["San Antonio"]
    points = [{"lat": hub["lat"], "lng": hub["lng"]}]
    for s in ordered_stores:
        points.append({"lat": num(s.get("lat")), "lng": num(s.get("lng"))})
    points.append({"lat": hub["lat"], "lng": hub["lng"]})

    racks = sum(num(s.get("expected_racks")) for s in ordered_stores)
    pieces = racks * PIECES_PER_RACK
    weight = sum(num(s.get("weight")) for s in ordered_stores)
    remaining = MAX_PAYLOAD - weight
    status = "OVER LIMIT" if weight > MAX_PAYLOAD else ("WARNING" if weight > WARNING_PAYLOAD else "SAFE")

    return {
        "hub": hub_name,
        "store_count": len(ordered_stores),
        "racks": round(racks, 2),
        "pieces": round(pieces, 2),
        "weight": round(weight, 2),
        "remaining_capacity": round(remaining, 2),
        "mileage": route_miles(points),
        "revenue": round(pieces * RATE_PER_PIECE, 2),
        "driver_pay": round(pieces * DRIVER_PAY_PER_PIECE, 2),
        "status": status
    }

def build_route_order(store_ids, mode):
    all_stores = read_json(STORES_FILE)
    selected = [s for s in all_stores if s.get("id") in store_ids and s.get("status") == "Unassigned"]
    if not selected:
        return None, [], None

    hub_name = selected[0].get("hub") if selected[0].get("hub") in HUBS else "San Antonio"

    if mode == "selection":
        order = {sid: i for i, sid in enumerate(store_ids)}
        ordered = sorted(selected, key=lambda s: order.get(s.get("id"), 999))
    else:
        ordered = order_nearest_from_hub(hub_name, selected)

    metrics = calculate_route_metrics(ordered, hub_name)
    return hub_name, ordered, metrics

def coords_for(city, state):
    city = clean(city)
    return CITY_COORDS.get(city, (30.2672, -97.7431))


def full_address_for_item(address, city, state, zip_code):
    parts = [clean(address), clean(city), clean(state), clean(zip_code)]
    return ", ".join([p for p in parts if p])

def geocode_address_azure(address):
    settings_data = read_json(SETTINGS_FILE)
    api_key = clean(settings_data.get("azure_maps_key"))
    address = clean(address)

    if not api_key or not address:
        return None, None, "No Azure Maps key or address"

    try:
        response = requests.get(
            "https://atlas.microsoft.com/search/address/json",
            params={
                "api-version": "1.0",
                "subscription-key": api_key,
                "query": address,
                "limit": 1,
                "countrySet": "US",
            },
            timeout=12
        )
        data = response.json()
        if data.get("results"):
            result = data["results"][0]
            pos = result.get("position") or {}
            lat = pos.get("lat")
            lng = pos.get("lon")
            freeform = (result.get("address") or {}).get("freeformAddress") or address
            if lat and lng:
                return lat, lng, freeform
        return None, None, f"Azure geocode failed: {data.get('error', {}).get('message') or data.get('summary', {}).get('query') or 'No result'}"
    except Exception as e:
        return None, None, f"Geocode error: {str(e)[:120]}"

def resolve_store_coordinates(address, city, state, zip_code):
    full_address = full_address_for_item(address, city, state, zip_code)
    lat, lng, note = geocode_address_azure(full_address)

    if lat and lng:
        return lat, lng, f"Azure Maps geocoded: {note}", full_address

    # Fallback only when geocoding is unavailable.
    lat, lng = coords_for(city, state)
    return lat, lng, f"City fallback: {note}", full_address

def assign_hub(lat, lng, dispatch_group=""):
    group = clean(dispatch_group).lower()
    if "houston" in group: return "Houston", "Dispatch Group"
    if "san antonio" in group or group == "sa" or "susatxus" in group: return "San Antonio", "Dispatch Group"
    if "dallas" in group or "irving" in group: return "Dallas", "Dispatch Group"

    nearest, nearest_miles = None, 999999
    for hub_name, hub in HUBS.items():
        miles = miles_between(lat, lng, hub["lat"], hub["lng"])
        if miles < nearest_miles:
            nearest, nearest_miles = hub_name, miles

    if nearest_miles <= 100:
        return nearest, f"Nearest Hub {round(nearest_miles, 1)} mi"
    return "Manual Review", f"Outside 100 mi ({round(nearest_miles, 1)} mi)"

def extract_pdf_text(path):
    reader = PdfReader(str(path))
    text = []
    for page in reader.pages:
        try:
            text.append(page.extract_text() or "")
        except Exception:
            pass
    return "\n".join(text)

def find_match(pattern, text, default=""):
    m = re.search(pattern, text, re.I | re.S)
    return clean(m.group(1)) if m else default

def find_corner_post_qty(text):
    text = text or ""
    patterns = [
        r'84"\s*Corner Post Only\s+R-CPH84\s+\d+\s+[\d,]+\s+([\d,]+)',
        r'84\s*"?\s*Corner Post Only.*?R-CPH84.*?([\d,]+)\s*(?:$|\n)',
        r'R-CPH84[^\n]*?([\d,]+)\s*(?:$|\n)',
    ]
    for pattern in patterns:
        value = num(find_match(pattern, text))
        if value:
            return value
    for line in text.splitlines():
        if "corner post" in line.lower() and ("84" in line or "R-CPH84" in line.upper()):
            nums = re.findall(r"\d[\d,]*", line.replace('"', " "))
            if nums:
                return num(nums[-1])
    return 0

def parse_city_state_zip(value):
    value = clean(value).replace("\n", " ")
    m = re.search(r"(.+?),\s*([A-Za-z]+)\s+(\d{5})", value)
    if not m:
        m = re.search(r"(.+?)\s+(Texas|TX)\s+(\d{5})", value, re.I)
    if m:
        city = clean(m.group(1))
        state = "TX" if m.group(2).lower() in ["texas", "tx"] else clean(m.group(2))
        zip_code = clean(m.group(3))
        return city, state, zip_code
    return value, "TX", ""

def parse_rms_pdf(path):
    text = extract_pdf_text(path)

    bol = find_match(r"Bill of Lading:\s*([0-9]+)", text)
    origin = find_match(r"Origin:\s*([A-Z0-9_-]+)", text)
    store_name = find_match(r"Origin:\s*[A-Z0-9_-]+\s*Name:\s*([^\n]+)", text) or find_match(r"Name:\s*([^\n]+)", text)
    address = find_match(r"Origin:.*?Address:\s*([^\n]+)", text)
    city_state_zip = find_match(r"Origin:.*?City/State/Zip:\s*([^\n]+)", text)
    destination = find_match(r"Destination:\s*([A-Z0-9_-]+)", text)
    city, state, zip_code = parse_city_state_zip(city_state_zip)

    corner_posts = find_corner_post_qty(text)
    drb40 = num(find_match(r'40"\s*DRB\s+R-DRB40\s+\d+\s+[\d,]+\s+([\d,]+)', text))
    drb48 = num(find_match(r'48"\s*DRB\s+R-DRB48\s+\d+\s+[\d,]+\s+([\d,]+)', text))
    wood_shelf = num(find_match(r'Wood Shelf\s+R-W4048\s+\d+\s+[\d,]+\s+([\d,]+)', text))
    assigned_date = find_match(r"Assigned Date\s*[:\s]*([0-9]{2}/[0-9]{2}/[0-9]{4})", text)
    due_date = find_match(r"Due Date(?: \(PDT\))?\s*[:\s]*([0-9]{2}/[0-9]{2}/[0-9]{4})", text)
    bol_weight = num(find_match(r"Bill of Lading Total Weight:\s*([\d,]+)", text) or find_match(r"Order Total Weight:\s*([\d,]+)", text))

    expected_racks = round(corner_posts / 4, 2) if corner_posts else 0
    lat, lng, geocode_status, full_address = resolve_store_coordinates(address, city, state, zip_code)
    hub, hub_reason = assign_hub(lat, lng, destination)

    item = {
        "id": str(uuid4()),
        "bol": bol,
        "origin": origin,
        "store_name": store_name,
        "address": address,
        "full_address": full_address,
        "geocode_status": geocode_status,
        "city": city,
        "state": state,
        "zip": zip_code,
        "lat": lat,
        "lng": lng,
        "hub": hub,
        "hub_reason": hub_reason,
        "dispatch_group": destination,
        "assigned_date": assigned_date,
        "due_date": due_date,
        "expected_racks": expected_racks,
        "corner_posts": corner_posts,
        "drb40": drb40,
        "drb48": drb48,
        "wood_shelf": wood_shelf,
        "weight": bol_weight or round(expected_racks * DEFAULT_RACK_WEIGHT, 2),
        "assigned_driver": "",
        "pdf_path": "",
        "created_at": datetime.now().isoformat(timespec="seconds")
    }
    item["review_reasons"] = essential_review_reasons(item)
    item["review_warnings"] = []
    if not origin:
        item["review_warnings"].append("Missing origin/store number")
    item["status"] = "Need Review" if item["review_reasons"] else "Unassigned"
    return item

def apply_material_counts(item, data):
    material_fields = {"corner_posts", "drb40", "drb48", "wood_shelf"}
    changed = False
    for key in material_fields:
        if key in data:
            item[key] = num(data.get(key))
            changed = True
    if "expected_racks" in data:
        item["expected_racks"] = num(data.get("expected_racks"))
    elif changed:
        item["expected_racks"] = round(num(item.get("corner_posts")) / 4, 2) if num(item.get("corner_posts")) else 0
    if "expected_pieces" in data:
        item["expected_pieces"] = num(data.get("expected_pieces"))
    elif changed:
        item["expected_pieces"] = num(item.get("expected_racks")) * PIECES_PER_RACK
    return item

def sync_route_stop_materials(store):
    routes = read_json(ROUTES_FILE)
    changed = False
    for route in routes:
        route_changed = False
        for stop in route.get("stops", []):
            if stop.get("id") != store.get("id"):
                continue
            for key in ("corner_posts", "drb40", "drb48", "wood_shelf", "expected_racks", "expected_pieces", "weight"):
                if key in store:
                    stop[key] = store.get(key)
            route_changed = True
        if route_changed and route.get("stops"):
            route["metrics"] = calculate_route_metrics(route["stops"], route.get("hub") or "San Antonio")
            changed = True
    if changed:
        write_json(ROUTES_FILE, routes)

def normalize_row(row):
    bol = clean(row.get("BOL #") or row.get("BOL") or row.get("BOL Number"))
    origin = clean(row.get("Origin") or row.get("Store") or row.get("Store Name"))
    store_name = clean(row.get("Store Name")) or origin or f"BOL {bol}"
    city = clean(row.get("City") or row.get("Origin City"))
    state = clean(row.get("State") or row.get("Origin State")) or "TX"
    address = clean(row.get("Origin Address") or row.get("Carrier Address") or row.get("Address"))
    dispatch_group = clean(row.get("Dispatch Group") or row.get("Route"))
    racks = num(row.get("Est Racks") or row.get("Expected Racks") or row.get("Racks"))

    zip_code = clean(row.get("Zip") or row.get("ZIP") or row.get("Origin Zip"))
    lat, lng, geocode_status, full_address = resolve_store_coordinates(address, city, state, zip_code)
    hub, hub_reason = assign_hub(lat, lng, dispatch_group)
    return {"id": str(uuid4()), "bol": bol, "origin": origin, "store_name": store_name, "address": address, "full_address": full_address, "geocode_status": geocode_status, "city": city, "state": state, "zip": zip_code, "lat": lat, "lng": lng, "hub": hub, "hub_reason": hub_reason, "dispatch_group": dispatch_group, "expected_racks": racks, "weight": round(racks * DEFAULT_RACK_WEIGHT, 2), "status": "Unassigned", "assigned_driver": "", "pdf_path": "", "created_at": datetime.now().isoformat(timespec="seconds")}

def parse_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "BOL_Intake" if "BOL_Intake" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        raw, has_data = {}, False
        for c, header in enumerate(headers, start=1):
            value = ws.cell(r, c).value
            if value not in [None, ""]: has_data = True
            raw[header] = value
        if has_data: rows.append(normalize_row(raw))
    return rows

def parse_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            rows.append(normalize_row(row))
    return rows

def import_rms_uploaded_files(files, source="RMS Import"):
    existing = read_json(STORES_FILE)
    existing_keys = {bol_duplicate_key(s) for s in existing if clean(s.get("bol")) or clean(s.get("origin"))}
    existing_bols = {clean(s.get("bol")) for s in existing if clean(s.get("bol"))}
    added, duplicates, review = 0, 0, 0
    errors = []

    for uploaded in files:
        filename = secure_filename(uploaded.filename or "")
        if not filename:
            continue
        temp_path = UPLOAD_DIR / f"{uuid4()}_{filename}"
        uploaded.save(temp_path)

        imported = []
        try:
            lower = filename.lower()
            if lower.endswith(".pdf"):
                item = parse_rms_pdf(temp_path)
                clean_name = f"BOL_{safe_part(item.get('bol'))}_{safe_part(item.get('origin'))}_{safe_part(item.get('store_name'))}_{safe_part(item.get('city'))}_{safe_part(item.get('state'))}.pdf"
                root = "Need_Review" if item["status"] == "Need Review" else "Imported"
                final_path = month_folder(root) / clean_name
                if final_path.exists():
                    final_path = final_path.with_name(final_path.stem + "_" + datetime.now().strftime("%H%M%S") + final_path.suffix)
                shutil.move(str(temp_path), str(final_path))
                item["pdf_path"] = str(final_path)
                imported = [item]
            elif lower.endswith(".xlsx"):
                imported = parse_xlsx(temp_path)
            elif lower.endswith(".csv"):
                imported = parse_csv(temp_path)
            else:
                errors.append(f"{filename}: unsupported file type")
        except Exception as exc:
            errors.append(f"{filename}: {str(exc)[:180]}")
        finally:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                pass

        for item in imported:
            if is_duplicate_bol(item, existing_keys, existing_bols):
                duplicates += 1
                continue
            existing.append(item)
            track_bol_key(item, existing_keys, existing_bols)
            added += 1
            if item.get("status") == "Need Review":
                review += 1

    write_json(STORES_FILE, existing)
    result = {
        "ok": True,
        "status": "IMPORT COMPLETE",
        "source": source,
        "added": added,
        "duplicates": duplicates,
        "need_review": review,
        "errors": errors[:20],
        "message": f"Import complete. Added {added}. Need Review {review}. Skipped duplicates {duplicates}."
    }
    audit(source, result)
    return result

def users_payload():
    if has_request_context() and hasattr(g, "_users_payload"):
        return g._users_payload
    data = read_json(USERS_FILE)
    if isinstance(data, dict) and "users" in data:
        if has_request_context():
            g._users_payload = data
        return data
    fallback = {"users": []}
    if has_request_context():
        g._users_payload = fallback
    return fallback

def save_users_payload(data):
    write_json(USERS_FILE, data)
    if has_request_context():
        g._users_payload = data
    if has_request_context() and hasattr(g, "_current_user"):
        delattr(g, "_current_user")

def current_user():
    if has_request_context() and hasattr(g, "_current_user"):
        return g._current_user
    username = session.get("username")
    if not username:
        if has_request_context():
            g._current_user = None
        return None
    for user in users_payload().get("users", []):
        if user.get("username") == username and user.get("active", True):
            if has_request_context():
                g._current_user = user
            return user
    if username == ADMIN_USERNAME:
        admin = {"id":"admin","username":ADMIN_USERNAME,"role":"Admin","assigned_cities":["All"],"active":True}
        if has_request_context():
            g._current_user = admin
        return admin
    if has_request_context():
        g._current_user = None
    return None

def is_admin():
    user = current_user()
    return bool(user and user.get("role") == "Admin")

def current_role():
    user = current_user()
    return clean(user.get("role") if user else "")

def is_operations_manager():
    return current_role() == "Operations Manager"

def can_dispatch():
    return current_role() in {"Admin", "Operations Manager", "Dispatcher"}

def user_allowed_city_values():
    user = current_user()
    if not user:
        return []
    cities = user.get("assigned_cities") or []
    if user.get("role") in {"Admin", "Operations Manager"} or "All" in cities:
        return ["All"]
    return cities

def store_city_allowed(store):
    allowed = user_allowed_city_values()
    if "All" in allowed:
        return True
    values = {clean(store.get("city")), clean(store.get("hub")), clean(store.get("dispatch_group")), clean(store.get("origin_city"))}
    return any(a in values for a in allowed)

def filter_stores_for_user(stores):
    return stores if is_admin() else [s for s in stores if store_city_allowed(s)]

def filter_routes_for_user(routes):
    if is_admin():
        return routes
    allowed = user_allowed_city_values()
    if "All" in allowed:
        return routes
    result = []
    for route in routes:
        vals = {clean(route.get("hub")), clean(route.get("city"))}
        for stop in route.get("stops", []):
            vals.add(clean(stop.get("city")))
            vals.add(clean(stop.get("hub")))
        if any(a in vals for a in allowed):
            result.append(route)
    return result

def filter_queue_for_user(queue):
    if is_admin():
        return queue
    allowed = user_allowed_city_values()
    if "All" in allowed:
        return queue
    return [q for q in queue if any(a in {clean(q.get("city")), clean(q.get("hub")), clean(q.get("dispatch_group")), clean(q.get("origin_city"))} for a in allowed)]

def map_settings_payload():
    settings_data = read_json(SETTINGS_FILE)
    return {
        "map_default_type": clean(settings_data.get("map_default_type")) or "satellite",
        "map_default_zoom": int(num(settings_data.get("map_default_zoom")) or 7),
        "map_board_default_open": bool(settings_data.get("map_board_default_open")),
        "map_live_refresh_seconds": int(num(settings_data.get("map_live_refresh_seconds")) or 30),
        "due_red_days": int(num(settings_data.get("due_red_days")) or 4),
        "due_amber_days": int(num(settings_data.get("due_amber_days")) or 7),
    }

def active_map_stores(stores):
    hidden_statuses = {"Completed", "RMS Closed"}
    hidden_rms = {"Closed in RMS", "Missing from RMS"}
    return [
        s for s in stores
        if (s.get("status") or "Unassigned") not in hidden_statuses
        and clean(s.get("rms_status")) not in hidden_rms
    ]

def today_iso():
    return datetime.now().date().isoformat()

def date_value(value):
    value = clean(value)
    if not value:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value[:10], fmt).date().isoformat()
        except Exception:
            pass
    return value[:10]

def item_touched_today(item):
    today = today_iso()
    for key in ("updated_at", "imported_at", "created_at", "completed_at", "last_seen", "last_seen_in_rms_at"):
        if date_value(item.get(key)) == today:
            return True
    return False

def essential_review_reasons(item):
    reasons = []
    if not clean(item.get("bol")):
        reasons.append("Missing BOL")
    if not (clean(item.get("store_name")) or clean(item.get("origin_name")) or clean(item.get("origin"))):
        reasons.append("Missing store/customer name")
    if not clean(item.get("city")):
        reasons.append("Missing city")
    if num(item.get("expected_racks")) <= 0:
        reasons.append('Missing 84" Corner Post quantity')
    if clean(item.get("hub")) == "Manual Review":
        reasons.append("Hub outside 100 miles")
    return reasons

def filtered_bols(stores):
    status_filter = clean(request.args.get("status"))
    imported_today = clean(request.args.get("imported")).lower() == "today"
    updated_today = clean(request.args.get("updated")).lower() == "today"
    completed_today = clean(request.args.get("today")).lower() in {"1", "true", "yes"}
    q = clean(request.args.get("q")).lower()

    rows = list(stores)
    if status_filter:
        rows = [
            s for s in rows
            if clean(s.get("status") or "Unassigned") == status_filter
            or clean(s.get("queue_status")) == status_filter
        ]
    if imported_today:
        rows = [s for s in rows if date_value(s.get("created_at") or s.get("imported_at")) == today_iso()]
    if updated_today:
        rows = [s for s in rows if item_touched_today(s)]
    if completed_today:
        rows = [s for s in rows if date_value(s.get("completed_at")) == today_iso()]
    if q:
        rows = [
            s for s in rows
            if q in " ".join(clean(s.get(k)).lower() for k in ("bol", "store_name", "origin", "city", "hub", "status"))
        ]
    return sorted(rows, key=lambda s: clean(s.get("updated_at") or s.get("created_at")), reverse=True)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        if not is_admin():
            return render_template("access_denied.html"), 403
        return f(*args, **kwargs)
    return decorated_function

def dispatch_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        if not can_dispatch():
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "Dispatch access required."}), 403
            return render_template("access_denied.html"), 403
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def enforce_login():
    path = request.path or "/"
    if path == "/" or path.startswith("/login") or path.startswith("/logout") or path.startswith("/static/") or path.startswith("/favicon"):
        return None
    if path == "/api/local-rms/import" and local_rms_import_authorized():
        return None
    if session.get("logged_in") and current_user():
        role = current_role()
        admin_only = (
            path.startswith("/users") or path.startswith("/settings") or
            path.startswith("/rms-sync") or path.startswith("/rms-import") or
            path.startswith("/rms-queue") or path.startswith("/api/rms/")
        )
        if admin_only and role != "Admin":
            if path.startswith("/api/"):
                return jsonify({"ok": False, "message": "Admin access required."}), 403
            return render_template("access_denied.html"), 403
        if role == "Driver":
            driver_allowed = (
                path.startswith("/driver") or path.startswith("/api/driver/") or
                path.startswith("/route-view/") or path.startswith("/api/route/")
            )
            if not driver_allowed:
                if path.startswith("/api/"):
                    return jsonify({"ok": False, "message": "Driver access is limited to assigned routes."}), 403
                return redirect("/driver")
        return None
    session.clear()
    if path.startswith("/api/"):
        return jsonify({"ok": False, "message": "Authentication required."}), 401
    return redirect(url_for("login", next=path))

def safe_next_url(candidate):
    """Allow only local absolute paths as post-login destinations."""
    candidate = clean(candidate)
    if candidate.startswith("/") and not candidate.startswith("//"):
        return candidate
    return "/dashboard"

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("logged_in") and current_user():
        return redirect(request.args.get("next") or "/dashboard")
    if session.get("logged_in"):
        session.clear()
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        next_url = safe_next_url(request.form.get("next") or request.args.get("next"))
        matched = None
        payload = users_payload()
        for user in payload.get("users", []):
            if user.get("username") == username and user.get("active", True):
                ok, needs_upgrade = verify_password(user.get("password"), password)
                if ok:
                    if needs_upgrade:
                        user["password"] = hash_password(password)
                        save_users_payload(payload)
                    matched = user
                    break
        if ADMIN_PASSWORD and not matched and username == ADMIN_USERNAME and secrets.compare_digest(password, ADMIN_PASSWORD):
            matched = {"username": ADMIN_USERNAME, "role":"Admin", "assigned_cities":["All"]}
        if matched:
            session.clear()
            session["logged_in"] = True
            session["username"] = matched.get("username")
            session["role"] = matched.get("role", "Dispatcher")
            session["assigned_cities"] = matched.get("assigned_cities", [])
            return redirect(next_url)
        error = "Invalid username or password."
    return render_template("login.html", error=error, next=request.args.get("next") or "/dashboard")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
def home():
    if session.get("logged_in"):
        return redirect("/dashboard")
    return redirect("/login")






def dashboard_metrics(stores=None, routes=None):
    stores = stores if stores is not None else filter_stores_for_user(read_json(STORES_FILE))
    routes = routes if routes is not None else filter_routes_for_user(read_json(ROUTES_FILE))
    statuses = ["Need Review", "Unassigned", "Assigned", "Dispatched", "Completed"]
    by_status = {status: 0 for status in statuses}
    for store in stores:
        status = store.get("status") or "Unassigned"
        by_status[status] = by_status.get(status, 0) + 1
    active = active_map_stores(stores)
    completed_today = [
        s for s in stores
        if (s.get("status") or "") == "Completed" and date_value(s.get("completed_at")) == today_iso()
    ]
    racks = round(sum(num(s.get("expected_racks")) for s in active), 1)
    weight = round(sum(num(s.get("weight")) for s in active), 1)
    pieces = round(racks * PIECES_PER_RACK, 1)
    open_routes = [r for r in routes if (r.get("status") or "Assigned") != "Completed"]
    recent_routes = sorted(routes, key=lambda r: r.get("created_at", ""), reverse=True)[:8]
    return {
        "by_status": by_status,
        "completed_today": len(completed_today),
        "active_stores": len(active),
        "open_routes": len(open_routes),
        "racks": racks,
        "weight": weight,
        "pieces": pieces,
        "revenue": round(pieces * RATE_PER_PIECE, 2),
        "driver_pay": round(pieces * DRIVER_PAY_PER_PIECE, 2),
        "remaining_capacity": round(MAX_PAYLOAD - weight, 1),
        "recent_routes": recent_routes,
    }

@app.route("/dashboard")
def dashboard():
    stores = filter_stores_for_user(read_json(STORES_FILE))
    routes = filter_routes_for_user(read_json(ROUTES_FILE))
    sync_history = read_json(SYNC_HISTORY_FILE)
    settings_data = read_json(SETTINGS_FILE)
    maps_key = clean(settings_data.get("azure_maps_key"))
    return render_template(
        "dashboard.html",
        metrics=dashboard_metrics(stores, routes),
        stores=active_map_stores(stores),
        routes=routes,
        hubs=HUBS,
        sync_history=sync_history[-5:] if isinstance(sync_history, list) else [],
        azure_maps_key=maps_key,
        map_settings=map_settings_payload(),
        can_view_financials=current_role() in {"Admin", "Operations Manager"},
    )

@app.route("/api/dashboard-live")
def api_dashboard_live():
    return jsonify({"ok": True, "metrics": dashboard_metrics()})

@app.route("/api/drivers")
@dispatch_required
def api_drivers():
    users = users_payload().get("users", [])
    drivers = []
    allowed_cities = set(user_allowed_city_values())
    for user in users:
        if not user.get("active", True):
            continue
        role = (user.get("role") or "").lower()
        if role == "driver":
            driver_cities = set(user.get("assigned_cities") or [])
            if current_role() == "Dispatcher" and "All" not in allowed_cities and not (allowed_cities & driver_cities):
                continue
            drivers.append({
                "name": user.get("display_name") or user.get("username"),
                "username": user.get("username"),
                "role": user.get("role", "Dispatcher"),
                "phone": user.get("phone", ""),
                "cities": user.get("assigned_cities", []),
            })
    return jsonify({"ok": True, "drivers": drivers})


@app.route("/drivers")
@dispatch_required
def drivers_directory():
    drivers = [u for u in users_payload().get("users", []) if (u.get("role") or "").lower() == "driver"]
    allowed_cities = set(user_allowed_city_values())
    if current_role() == "Dispatcher" and "All" not in allowed_cities:
        drivers = [u for u in drivers if allowed_cities & set(u.get("assigned_cities") or [])]
    routes = filter_routes_for_user(read_json(ROUTES_FILE))
    driver_phones = {}
    for driver in drivers:
        phone = clean(driver.get("phone"))
        if phone:
            driver_phones[clean(driver.get("username"))] = phone
            driver_phones[clean(driver.get("display_name"))] = phone
    route_counts = {}
    for route in routes:
        driver_name = clean(route.get("driver"))
        if driver_name:
            route_counts[driver_name] = route_counts.get(driver_name, 0) + 1
            if not clean(route.get("driver_phone")):
                route["driver_phone"] = driver_phones.get(driver_name, "")
    return render_template(
        "drivers.html", drivers=drivers, routes=routes,
        route_counts=route_counts, city_options=["All", "San Antonio", "Houston", "Dallas"],
        can_manage=is_admin()
    )


@app.route("/drivers/create", methods=["POST"])
@admin_required
def drivers_create():
    username = clean(request.form.get("username"))
    display_name = clean(request.form.get("display_name"))
    password = request.form.get("password") or ""
    phone = clean(request.form.get("phone"))
    assigned_cities = request.form.getlist("assigned_cities") or ["San Antonio"]
    if "All" in assigned_cities:
        assigned_cities = ["San Antonio", "Houston", "Dallas"]
    data = users_payload()
    users = data.get("users", [])
    if not username or not display_name or not password or not phone:
        return redirect("/drivers?error=missing")
    if any(clean(u.get("username")) == username for u in users):
        return redirect("/drivers?error=duplicate")
    users.append({
        "id": str(uuid4()), "username": username, "display_name": display_name,
        "password": hash_password(password), "phone": phone, "role": "Driver",
        "assigned_cities": assigned_cities, "active": True,
        "created_at": datetime.now().isoformat(timespec="seconds")
    })
    data["users"] = users
    save_users_payload(data)
    audit("Create Driver", {"username": username, "display_name": display_name, "phone": phone, "assigned_cities": assigned_cities})
    return redirect("/drivers?created=1")


@app.route("/drivers/update/<driver_id>", methods=["POST"])
@admin_required
def drivers_update(driver_id):
    data = users_payload()
    updated = None
    for user in data.get("users", []):
        if (user.get("id") == driver_id or user.get("username") == driver_id) and (user.get("role") or "").lower() == "driver":
            user["display_name"] = clean(request.form.get("display_name")) or user.get("display_name") or user.get("username")
            user["phone"] = clean(request.form.get("phone")) or user.get("phone", "")
            user["assigned_cities"] = request.form.getlist("assigned_cities") or user.get("assigned_cities", ["San Antonio"])
            user["active"] = bool(request.form.get("active"))
            password = request.form.get("password") or ""
            if password:
                user["password"] = hash_password(password)
            user["updated_at"] = datetime.now().isoformat(timespec="seconds")
            updated = user
            break
    save_users_payload(data)
    if updated:
        audit("Update Driver", {"username": updated.get("username"), "active": updated.get("active"), "assigned_cities": updated.get("assigned_cities")})
    return redirect("/drivers")

@app.route("/drivers/delete/<driver_id>", methods=["POST"])
@admin_required
def drivers_delete(driver_id):
    data = users_payload()
    users = data.get("users", [])
    target = next(
        (
            u for u in users
            if (u.get("id") == driver_id or u.get("username") == driver_id)
            and (u.get("role") or "").lower() == "driver"
        ),
        None,
    )
    if not target:
        audit("Delete Driver Failed", {"driver_id": driver_id, "reason": "not found"})
        return redirect("/drivers?error=notfound")

    data["users"] = [
        u for u in users
        if not (u.get("id") == target.get("id") or u.get("username") == target.get("username"))
    ]
    save_users_payload(data)
    audit("Delete Driver", {
        "username": target.get("username"),
        "display_name": target.get("display_name"),
        "phone": target.get("phone"),
    })
    return redirect("/drivers?deleted=1")

@app.route("/api/dispatch-map-debug")
def api_dispatch_map_debug():
    stores = read_json(STORES_FILE)
    return jsonify({
        "total_stores": len(stores),
        "unassigned": len([s for s in stores if s.get("status") == "Unassigned"]),
        "need_review": len([s for s in stores if s.get("status") == "Need Review"]),
        "assigned": len([s for s in stores if s.get("status") == "Assigned"]),
        "completed": len([s for s in stores if s.get("status") == "Completed"]),
        "map_eligible": [
            {
                "bol": s.get("bol"),
                "status": s.get("status"),
                "lat": s.get("lat"),
                "lng": s.get("lng"),
                "city": s.get("city"),
                "store_name": s.get("store_name"),
                "racks": s.get("expected_racks"),
                "address": s.get("address"),
                "full_address": s.get("full_address"),
                "geocode_status": s.get("geocode_status")
            }
            for s in stores if s.get("status") == "Unassigned"
        ]
    })


@app.route("/api/geocode-stores", methods=["POST"])
def api_geocode_stores():
    stores = read_json(STORES_FILE)
    updated = 0
    failed = 0

    for store in stores:
        address = store.get("address") or store.get("origin_address") or ""
        city = store.get("city") or store.get("origin_city") or ""
        state = store.get("state") or store.get("origin_state") or "TX"
        zip_code = store.get("zip") or store.get("origin_zip") or ""

        if not address or not city:
            failed += 1
            store["geocode_status"] = "Missing address or city"
            continue

        lat, lng, geocode_status, full_address = resolve_store_coordinates(address, city, state, zip_code)
        store["lat"] = lat
        store["lng"] = lng
        store["full_address"] = full_address
        store["geocode_status"] = geocode_status

        if "Azure Maps geocoded" in geocode_status:
            updated += 1
        else:
            failed += 1

    write_json(STORES_FILE, stores)
    audit("Geocode Stores", {"updated": updated, "failed": failed})

    return jsonify({"ok": True, "updated": updated, "failed": failed})

@app.route("/dispatch-map")
@dispatch_required
def dispatch_map():
    stores = active_map_stores(filter_stores_for_user(read_json(STORES_FILE)))
    status_filter = clean(request.args.get("status"))
    if status_filter:
        stores = [s for s in stores if clean(s.get("status") or "Unassigned") == status_filter]
    settings_data = read_json(SETTINGS_FILE)
    maps_key = clean(settings_data.get("azure_maps_key"))
    return render_template(
        "dispatch_map.html", stores=stores, hubs=HUBS,
        max_payload=MAX_PAYLOAD, azure_maps_key=maps_key,
        map_settings=map_settings_payload(),
        can_view_financials=current_role() in {"Admin", "Operations Manager"}
    )

@app.route("/route-builder")
def route_builder():
    routes = filter_routes_for_user(read_json(ROUTES_FILE))
    status_filter = clean(request.args.get("status"))
    if status_filter:
        routes = [r for r in routes if clean(r.get("status") or "Assigned") == status_filter]
    return render_template("route_builder.html", routes=routes)

@app.route("/rms-import", methods=["GET", "POST"])
def rms_import():
    if request.method == "POST":
        files = request.files.getlist("rms_file")
        if not files or files[0].filename == "":
            return render_template("rms_import.html", message="Choose RMS PDF, Excel, or CSV files first.")

        result = import_rms_uploaded_files(files, source="RMS Import")
        extra = f" Errors: {'; '.join(result['errors'][:3])}" if result.get("errors") else ""
        return render_template("rms_import.html", message=result["message"] + extra)

    return render_template("rms_import.html", message="")

@app.route("/api/local-rms/import", methods=["POST"])
def api_local_rms_import():
    if not clean(os.environ.get("LOCAL_RMS_IMPORT_TOKEN")):
        return jsonify({
            "ok": False,
            "message": "LOCAL_RMS_IMPORT_TOKEN is not configured in Azure App Service settings."
        }), 503
    if not local_rms_import_authorized():
        return jsonify({"ok": False, "message": "Invalid or missing local RMS import token."}), 401

    files = request.files.getlist("rms_file") or request.files.getlist("files")
    if not files:
        return jsonify({"ok": False, "message": "Attach PDF, Excel, or CSV files as rms_file."}), 400

    result = import_rms_uploaded_files(files, source="Local RMS Import")
    return jsonify(result)

@app.route("/need-review")
def need_review():
    stores = [s for s in filter_stores_for_user(read_json(STORES_FILE)) if s.get("status") == "Need Review"]
    return render_template("need_review.html", stores=stores, hubs=HUBS)

@app.route("/all-bols")
@dispatch_required
def all_bols():
    stores = filter_stores_for_user(read_json(STORES_FILE))
    queue = filter_queue_for_user(read_json(RMS_QUEUE_FILE))
    rows_by_bol = {}
    for store in stores:
        bol = clean(store.get("bol")) or store.get("id")
        store["_source"] = "store"
        store["_edit_id"] = store.get("id")
        rows_by_bol[bol] = store
    for q in queue:
        bol = clean(q.get("bol")) or q.get("id")
        if bol in rows_by_bol:
            rows_by_bol[bol]["queue_status"] = q.get("queue_status", rows_by_bol[bol].get("queue_status", ""))
            rows_by_bol[bol]["last_seen"] = q.get("last_seen", rows_by_bol[bol].get("last_seen", ""))
            continue
        q["_source"] = "queue"
        q["_edit_id"] = q.get("id") or bol
        q["status"] = q.get("queue_status") or "New"
        rows_by_bol[bol] = q
    rows = filtered_bols(list(rows_by_bol.values()))
    return render_template(
        "all_bols.html",
        stores=rows,
        hubs=HUBS,
        status_options=["New", "Imported", "Need Review", "Unassigned", "Assigned", "Dispatched", "Completed", "Ignored"],
        current_status=clean(request.args.get("status")),
        q=clean(request.args.get("q")),
    )

@app.route("/api/approve-review", methods=["POST"])
def api_approve_review():
    data = request.get_json(force=True)
    store_id = data.get("store_id")
    hub = data.get("hub")
    stores = read_json(STORES_FILE)
    updated = None
    for store in stores:
        if store.get("id") == store_id:
            store["hub"] = hub or store.get("hub")
            store["status"] = "Unassigned"
            store["review_reasons"] = []
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Imported")
            updated = store
            break
    write_json(STORES_FILE, stores)
    audit("Approve Need Review", {"store_id": store_id, "hub": hub})
    return jsonify({"ok": True, "store": updated})

@app.route("/api/delete-bol", methods=["POST"])
@dispatch_required
def api_delete_bol():
    data = request.get_json(force=True)
    store_id = clean(data.get("store_id"))
    bol = clean(data.get("bol"))

    stores = read_json(STORES_FILE)
    queue = read_json(RMS_QUEUE_FILE)
    routes = read_json(ROUTES_FILE)
    removed_stores = []
    kept_stores = []

    for store in stores:
        matches_id = store_id and store.get("id") == store_id
        matches_bol = bol and clean(store.get("bol")) == bol
        if matches_id or matches_bol:
            removed_stores.append(store)
        else:
            kept_stores.append(store)

    removed_store_ids = {clean(store.get("id")) for store in removed_stores if clean(store.get("id"))}
    target_bol = bol or next((clean(store.get("bol")) for store in removed_stores if clean(store.get("bol"))), "")
    deleted_files = []
    failed_file_deletes = []
    for store in removed_stores:
        cleanup = delete_saved_bol_files(store)
        deleted_files.extend(cleanup.get("deleted", []))
        failed_file_deletes.extend(cleanup.get("failed", []))

    kept_queue = []
    removed_queue_items = []
    for q in queue:
        matches_id = store_id and clean(q.get("id")) == store_id
        matches_bol = target_bol and clean(q.get("bol")) == target_bol
        if matches_id or matches_bol:
            removed_queue_items.append(q)
            if not target_bol:
                target_bol = clean(q.get("bol"))
        else:
            kept_queue.append(q)

    for q in removed_queue_items:
        cleanup = delete_saved_bol_files(q)
        deleted_files.extend(cleanup.get("deleted", []))
        failed_file_deletes.extend(cleanup.get("failed", []))

    if not removed_stores and not removed_queue_items:
        return jsonify({"ok": False, "message": "BOL not found in All BOLs, Need Review, store records, or RMS queue."}), 404

    removed_route_refs = 0
    for route in routes:
        before_ids = list(route.get("store_ids") or [])
        if before_ids:
            route["store_ids"] = [sid for sid in before_ids if clean(sid) not in removed_store_ids]
            removed_route_refs += len(before_ids) - len(route["store_ids"])

        before_stops = list(route.get("stops") or [])
        if before_stops:
            kept_stops = []
            for stop in before_stops:
                stop_id = clean(stop.get("id"))
                stop_bol = clean(stop.get("bol"))
                if (stop_id and stop_id in removed_store_ids) or (target_bol and stop_bol == target_bol):
                    removed_route_refs += 1
                else:
                    kept_stops.append(stop)
            route["stops"] = kept_stops

    write_json(STORES_FILE, kept_stores)
    write_json(RMS_QUEUE_FILE, kept_queue)
    write_json(ROUTES_FILE, routes)
    audit("Delete BOL For Regrab", {
        "store_id": store_id,
        "bol": target_bol,
        "removed_stores": len(removed_stores),
        "removed_queue": len(removed_queue_items),
        "removed_route_refs": removed_route_refs,
        "deleted_files": deleted_files,
        "failed_file_deletes": failed_file_deletes,
    })
    message = f"Deleted BOL {target_bol or store_id}. It can be grabbed fresh on the next RMS scan."
    if failed_file_deletes:
        message += f" {len(failed_file_deletes)} saved file(s) could not be removed, but the EOMS record was deleted."
    return jsonify({
        "ok": True,
        "message": message,
        "bol": target_bol,
        "deleted_files": deleted_files,
        "failed_file_deletes": failed_file_deletes,
        "removed_stores": len(removed_stores),
        "removed_queue": len(removed_queue_items),
        "removed_route_refs": removed_route_refs,
    })



def rms_login_with_playwright(headless=True):
    settings_data = read_json(SETTINGS_FILE)

    username = clean(settings_data.get("rms_username"))
    password = clean(settings_data.get("rms_password"))
    login_url = normalized_rms_login_url(settings_data.get("rms_login_url"))
    bol_url = normalized_rms_bol_list_url(settings_data.get("rms_bol_url"))

    if not username or not password:
        return {
            "ok": False,
            "status": "MISSING CREDENTIALS",
            "message": "Save RMS username and password in Settings first.",
            "bol_count": 0,
            "sample_bols": []
        }

    with sync_playwright() as p:
        browser, context = launch_rms_browser_context(p, headless=headless)
        page = context.new_page()
        try:
            login_to_rms(page, login_url, username, password)

            # Required RMS flow step 2:
            #   https://rms.reusability.com/bills-of-lading
            bol_url = normalized_rms_bol_list_url(bol_url)
            response = page.goto(bol_url, wait_until="networkidle", timeout=60000)
            assert_rms_accessible(page, response, "RMS BOL list")
            if is_rms_login_page(page):
                raise RuntimeError(
                    "RMS redirected back to login when opening the BOL list. "
                    "Verify RMS credentials and that the account can access Bills of Lading. "
                    + page_excerpt(page)
                )

            # Detect BOL links across all pages
            bol_links = collect_bol_links_from_all_pages(page)

            # Save diagnostic screenshot for troubleshooting
            diag_dir = BASE_DIR / "diagnostics"
            diag_dir.mkdir(exist_ok=True)
            screenshot_path = diag_dir / f"rms_sync_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            page.screenshot(path=str(screenshot_path), full_page=True)

            rms_debug_hold(page)
            close_rms_browser(browser, context)

            return {
                "ok": True,
                "status": "LOGIN SUCCESS",
                "message": f"RMS login successful. Found {len(bol_links)} BOL links on the Bills of Lading page.",
                "bol_count": len(bol_links),
                "sample_bols": bol_links[:10],
                "screenshot": str(screenshot_path)
            }

        except PlaywrightTimeoutError as e:
            close_rms_browser(browser, context)
            return {
                "ok": False,
                "status": "TIMEOUT",
                "message": f"RMS page timed out: {str(e)[:200]}",
                "bol_count": 0,
                "sample_bols": []
            }
        except Exception as e:
            close_rms_browser(browser, context)
            return {
                "ok": False,
                "status": "ERROR",
                "message": f"RMS login automation error: {str(e)[:300]}",
                "bol_count": 0,
                "sample_bols": []
            }




def extract_printable_bol_from_text(text, source_url=""):
    raw = text or ""
    lines = [clean(x) for x in raw.splitlines() if clean(x)]
    joined = "\n".join(lines)

    bol = find_match(r"Bill of Lading[:\s#-]*([0-9]+)", joined)
    if not bol:
        bol = find_match(r"Bill of Lading\s*-\s*([0-9]+)", joined)
    if not bol:
        bol = find_match(r"/bills-of-lading/([0-9]+)/print", source_url)

    origin = find_match(r"Origin[:\s]+([A-Z0-9_-]+)", joined)
    destination = find_match(r"Destination[:\s]+([A-Z0-9_-]+)", joined)

    def capture_section(start_label, stop_labels):
        start_idx = None
        for i, line in enumerate(lines):
            if re.search(rf"^{re.escape(start_label)}", line, re.I):
                start_idx = i
                break
        if start_idx is None:
            return ""

        captured = []
        for line in lines[start_idx:start_idx + 18]:
            if captured and any(re.search(rf"^{re.escape(stop)}", line, re.I) for stop in stop_labels):
                break
            captured.append(line)
        return "\n".join(captured)

    origin_section = capture_section("Origin", ["Destination", "Carrier", "Bill of Lading", "Arrival Time"])
    destination_section = capture_section("Destination", ["Carrier", "Bill of Lading", "Arrival Time"])

    def field_from(section, label):
        m = re.search(rf"{label}:\s*([^\n]+)", section, re.I)
        return clean(m.group(1)) if m else ""

    origin_name = field_from(origin_section, "Name")
    origin_address = field_from(origin_section, "Address")
    origin_city_line = field_from(origin_section, "City/State/Zip")
    origin_contact = field_from(origin_section, "Contact")

    dest_name = field_from(destination_section, "Name")
    dest_address = field_from(destination_section, "Address")
    dest_city_line = field_from(destination_section, "City/State/Zip")
    dest_contact = field_from(destination_section, "Contact")

    city, state, zip_code = parse_city_state_zip(origin_city_line)
    if not city:
        city, state, zip_code = parse_city_state_zip(dest_city_line)

    # Store fields use origin/pickup side.
    store_name = origin_name
    address = origin_address
    contact = origin_contact

    # Fallbacks from any visible label.
    if not store_name:
        store_name = find_match(r"Name:\s*([^\n]+)", joined)
    if not address:
        address = find_match(r"Address:\s*([^\n]+)", joined)
    if not city:
        tx_line = ""
        for line in lines:
            if re.search(r",?\s*TX\s+\d{5}|Texas\s+\d{5}", line, re.I):
                tx_line = line
                break
        if tx_line:
            city, state, zip_code = parse_city_state_zip(tx_line)

    def item_qty(label):
        patterns = [
            rf'{re.escape(label)}\s+R-[A-Z0-9]+\s+[\d,]+\s+[\d,]+\s+([\d,]+)',
            rf'{re.escape(label)}.*?R-[A-Z0-9]+.*?(\d+)\s*$',
            rf'{re.escape(label)}.*?Qty[:\s]+([\d,]+)'
        ]
        for pat in patterns:
            m = re.search(pat, joined, re.I | re.M)
            if m:
                return num(m.group(1))

        # Table fallback: find the item line and use the last number on the same row.
        for line in lines:
            if label.lower() in line.lower():
                nums = re.findall(r"\d+", line.replace(",", ""))
                if nums:
                    return num(nums[-1])
        return 0

    corner_posts = item_qty('84" Corner Post Only') or item_qty("84 Corner Post Only") or find_corner_post_qty(joined)
    drb40 = item_qty('40" DRB')
    drb48 = item_qty('48" DRB')
    wood_shelf = item_qty('Wood Shelf')


    assigned_date = find_match(r"Assigned Date\s*[:\s]*([0-9]{2}/[0-9]{2}/[0-9]{4})", joined)
    due_date = find_match(r"Due Date(?: \(PDT\))?\s*[:\s]*([0-9]{2}/[0-9]{2}/[0-9]{4})", joined)

    bol_weight = num(
        find_match(r"Bill of Lading Total Weight:\s*([\d,]+)", joined)
        or find_match(r"Order Total Weight:\s*([\d,]+)", joined)
    )

    expected_racks = round(corner_posts / 4, 2) if corner_posts else 0
    lat, lng, geocode_status, full_address = resolve_store_coordinates(address, city, state, zip_code)
    hub, hub_reason = assign_hub(lat, lng, destination)

    item = {
        "id": str(uuid4()),
        "bol": bol,
        "origin": origin,
        "store_name": store_name,
        "address": address,
        "full_address": full_address,
        "geocode_status": geocode_status,
        "city": city,
        "state": state,
        "zip": zip_code,
        "contact": contact,
        "origin_name": origin_name,
        "origin_address": origin_address,
        "origin_city": city,
        "origin_state": state,
        "origin_zip": zip_code,
        "origin_contact": origin_contact,
        "destination": destination,
        "destination_name": dest_name,
        "destination_address": dest_address,
        "destination_city_state_zip": dest_city_line,
        "destination_contact": dest_contact,
        "lat": lat,
        "lng": lng,
        "hub": hub,
        "hub_reason": hub_reason,
        "dispatch_group": destination,
        "assigned_date": assigned_date,
        "due_date": due_date,
        "expected_racks": expected_racks,
        "corner_posts": corner_posts,
        "drb40": drb40,
        "drb48": drb48,
        "wood_shelf": wood_shelf,
        "weight": bol_weight or round(expected_racks * DEFAULT_RACK_WEIGHT, 2),
        "status": "Unassigned",
        "review_reasons": [],
        "assigned_driver": "",
        "pdf_path": "",
        "rms_url": source_url,
        "created_at": datetime.now().isoformat(timespec="seconds")
    }
    item["review_reasons"] = essential_review_reasons(item)
    item["review_warnings"] = []
    if not origin:
        item["review_warnings"].append("Missing origin/store number")
    item["status"] = "Need Review" if item["review_reasons"] else "Unassigned"
    return item

def save_printable_snapshot(item, html):
    root = "Need_Review" if item.get("status") == "Need Review" else "Imported"
    filename = f"BOL_{safe_part(item.get('bol'))}_{safe_part(item.get('origin'))}_{safe_part(item.get('store_name'))}_{safe_part(item.get('city'))}_{safe_part(item.get('state'))}.html"
    final_path = month_folder(root) / filename
    final_path.write_text(html, encoding="utf-8", errors="ignore")
    return str(final_path)

def save_printable_pdf(page, item, html=None):
    root = "Need_Review" if item.get("status") == "Need Review" else "Imported"
    filename = f"BOL_{safe_part(item.get('bol'))}_{safe_part(item.get('origin'))}_{safe_part(item.get('store_name'))}_{safe_part(item.get('city'))}_{safe_part(item.get('state'))}.pdf"
    final_path = month_folder(root) / filename
    if final_path.exists():
        final_path = final_path.with_name(final_path.stem + "_" + datetime.now().strftime("%H%M%S") + final_path.suffix)
    try:
        page.emulate_media(media="print")
        page.pdf(path=str(final_path), format="Letter", print_background=True, margin={"top": "0.25in", "right": "0.25in", "bottom": "0.25in", "left": "0.25in"})
        return str(final_path)
    except Exception:
        if html is None:
            html = page.content()
        return save_printable_snapshot(item, html)


def is_rms_wait_page(text, html=""):
    body = clean(text).lower()
    page = (html or "").lower()
    if body in {"please wait", "please wait.", "please wait..."}:
        return True
    if "please wait" in body and len(body) < 250:
        return True
    if "transformation" in page and "please wait" in page and len(page) < 5000:
        return True
    return False

def wait_for_printable_bol_content(page, bol_number, timeout_seconds=25):
    deadline = time.time() + timeout_seconds
    last_text = ""
    last_html = ""
    while time.time() < deadline:
        try:
            page.wait_for_load_state("networkidle", timeout=3000)
        except Exception:
            pass
        try:
            last_text = page.locator("body").inner_text(timeout=3000)
            last_html = page.content()
            if not is_rms_wait_page(last_text, last_html):
                bol = clean(bol_number)
                text_l = last_text.lower()
                if not bol or bol in last_text or "bill of lading" in text_l or "origin" in text_l:
                    return last_text, last_html
        except Exception:
            pass
        page.wait_for_timeout(500)
    return last_text, last_html

def find_saved_bol_path(identifier):
    target = clean(identifier)
    stores = read_json(STORES_FILE)
    queue = read_json(RMS_QUEUE_FILE)

    for item in stores + queue:
        if item.get("id") == target or clean(item.get("bol")) == target:
            for key in ("pdf_path", "printable_path"):
                path = clean(item.get(key))
                if path and Path(path).exists():
                    return Path(path)

    bol = safe_part(target)
    matches = list(BOL_DIR.rglob(f"*{bol}*.pdf")) + list(BOL_DIR.rglob(f"*{bol}*.html"))
    matches = [p for p in matches if p.exists()]
    if matches:
        matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return matches[0]
    return None


def extract_date_from_text(text):
    text = clean(text)
    patterns = [
        r"\b(\d{1,2}/\d{1,2}/\d{4})\b",
        r"\b(\d{4}-\d{2}-\d{2})\b"
    ]
    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            return clean(m.group(1))
    return ""

def normalize_due_date(value):
    value = clean(value)
    if not value:
        return ""
    try:
        if re.match(r"^\d{4}-\d{2}-\d{2}$", value):
            return datetime.strptime(value, "%Y-%m-%d").strftime("%m/%d/%Y")
        if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", value):
            return datetime.strptime(value, "%m/%d/%Y").strftime("%m/%d/%Y")
    except Exception:
        pass
    return value

RMS_LOGIN_URL = "https://rms.reusability.com/login"
RMS_BOL_LIST_URL = "https://rms.reusability.com/bills-of-lading"

def rms_print_url(bol_number):
    bol = clean(bol_number)
    return f"https://rms.reusability.com/bills-of-lading/{bol}/print"

def normalized_rms_login_url(value=""):
    # RMS must start here. Keep this fixed unless RMS changes their public route.
    return RMS_LOGIN_URL

def normalized_rms_bol_list_url(value=""):
    # After login, EOMS must open the Bills of Lading list here.
    return RMS_BOL_LIST_URL

RMS_USERNAME_SELECTOR = (
    'input[name="username"], input[name="email"], input[type="email"], '
    'input[id*="username" i], input[id*="email" i], '
    'input[placeholder*="Username" i], input[placeholder*="Email" i], '
    'input[autocomplete="username"], input:not([type]), input[type="text"]'
)
RMS_PASSWORD_SELECTOR = 'input[type="password"], input[name*="password" i], input[id*="password" i]'
RMS_BROWSER_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)

class RMSAccessError(RuntimeError):
    pass

def new_rms_context(browser):
    return browser.new_context(
        user_agent=RMS_BROWSER_USER_AGENT,
        viewport={"width": 1366, "height": 768},
        locale="en-US",
        timezone_id="America/Chicago",
        extra_http_headers={
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        }
    )

def rms_blocked_result(message, action="RMS Import"):
    return {
        "ok": False,
        "status": "RMS BLOCKED / 403",
        "message": (
            "The RMS login URL is correct: https://rms.reusability.com/login. "
            "Correct RMS flow is login → https://rms.reusability.com/bills-of-lading → https://rms.reusability.com/bills-of-lading/<BOL>/print. "
            "RMS returned 403 Forbidden before the login form loaded, so EOMS never got a chance to enter the username/password. "
            "This is a server/network security block, not a bad password or wrong URL. "
            "Fix: run EOMS from a computer/network that can manually open RMS, use RUN_LOCAL_EOMS.bat with RMS_HEADLESS=0 for headed mode, "
            "or ask RMS/Reusability to allow-list the Azure App Service outbound IP. "
            "Until then, use RMS PDF/CSV/XLSX Import in EOMS. Details: " + clean(message)[:900]
        ),
        "found": 0,
        "imported": 0,
        "updated": 0,
        "skipped": 0,
        "need_review": 0,
        "failed": 1,
        "fallback": "/rms-import",
        "action": action,
    }

def page_excerpt(page):
    try:
        title = clean(page.title())
    except Exception:
        title = ""
    try:
        body = clean(page.locator("body").inner_text(timeout=5000))
    except Exception:
        body = ""
    return f"url={page.url} title={title} body={body[:350]}"

def assert_rms_accessible(page, response=None, label="RMS page"):
    status = None
    try:
        status = response.status if response else None
    except Exception:
        status = None
    excerpt = page_excerpt(page)
    lowered = excerpt.lower()
    if status in {401, 403} or "title=403 forbidden" in lowered or "body=403 forbidden" in lowered:
        raise RMSAccessError(
            f"{label} returned {status or '403 Forbidden'}. The RMS login URL is correct, but RMS blocked this server/browser before the login form loaded. "
            "This normally means the current Azure/server IP, VPN, or automation browser is not allowed by RMS security rules. "
            "Open RMS from the same server/network, check RMS/Reusability IP allow-list/security rules, or run EOMS locally in headed mode with RMS_HEADLESS=0. "
            + excerpt
        )
    if status and status >= 500:
        raise RMSAccessError(f"{label} returned HTTP {status}. " + excerpt)

def is_rms_login_page(page):
    try:
        url = clean(page.url).lower()
    except Exception:
        url = ""
    try:
        title = clean(page.title()).lower()
    except Exception:
        title = ""
    try:
        body = clean(page.locator("body").inner_text(timeout=5000)).lower()
    except Exception:
        body = ""
    return (
        "/login" in url or
        "login" in title or
        ("sign in" in body and "username" in body and "password" in body)
    )

def open_rms_bol_list_with_manual_login(page, bol_url=None, wait_seconds=None):
    bol_url = normalized_rms_bol_list_url(bol_url or RMS_BOL_LIST_URL)
    if wait_seconds is None:
        try:
            wait_seconds = int(os.environ.get("RMS_MANUAL_LOGIN_TIMEOUT_SECONDS", "600") or 600)
        except Exception:
            wait_seconds = 600

    response = page.goto(bol_url, wait_until="domcontentloaded", timeout=90000)
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass
    assert_rms_accessible(page, response, "RMS BOL list")

    deadline = time.time() + max(10, wait_seconds)
    last_excerpt = page_excerpt(page)
    last_url = clean(page.url)
    retry_at = 0
    while time.time() < deadline:
        try:
            current = clean(page.url)
            current_l = current.lower()
            last_excerpt = page_excerpt(page)
            if "bills-of-lading" in current_l and not is_rms_login_page(page):
                return True

            # After the user completes RMS login, RMS may leave the tab on a
            # home/menu page. Keep nudging that same tab back to the BOL list.
            if current != last_url:
                last_url = current
                retry_at = 0
            if not is_rms_login_page(page) and time.time() >= retry_at:
                retry_at = time.time() + 5
                try:
                    page.goto(bol_url, wait_until="domcontentloaded", timeout=30000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=10000)
                    except Exception:
                        pass
                except Exception:
                    pass
        except Exception:
            pass
        page.wait_for_timeout(1000)

    raise RuntimeError(
        "RMS did not reach the Bills of Lading list. Keep the RMS browser open, log in manually, "
        "then open Orders & BOLs / Bills of Lading before the Auto Grab timer expires. "
        "Last page: " + last_excerpt
    )

def fill_first_visible(page, selector, value, label, timeout_ms=90000):
    deadline = time.time() + (timeout_ms / 1000)
    last_error = ""
    while time.time() < deadline:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 25)
            for idx in range(count):
                field = locator.nth(idx)
                try:
                    if field.is_visible(timeout=1000) and field.is_enabled(timeout=1000):
                        field.click(timeout=5000)
                        try:
                            field.press("Control+A", timeout=3000)
                            field.press("Backspace", timeout=3000)
                        except Exception:
                            field.fill("", timeout=5000)
                        field.type(value, delay=35, timeout=30000)
                        page.wait_for_timeout(250)
                        entered_length = field.evaluate("(el) => (el.value || '').length", timeout=5000)
                        if entered_length > 0:
                            return True
                        field.fill(value, timeout=15000)
                        entered_length = field.evaluate("(el) => (el.value || '').length", timeout=5000)
                        if entered_length > 0:
                            return True
                        last_error = f"{label} field stayed empty after typing"
                except Exception as exc:
                    last_error = str(exc)[:160]
        except Exception as exc:
            last_error = str(exc)[:160]
        page.wait_for_timeout(750)
    raise PlaywrightTimeoutError(
        f"Timed out waiting for RMS {label} field. {page_excerpt(page)} last_error={last_error}"
    )

def click_first_visible(page, selector, label, timeout_ms=45000):
    deadline = time.time() + (timeout_ms / 1000)
    last_error = ""
    while time.time() < deadline:
        try:
            locator = page.locator(selector)
            count = min(locator.count(), 25)
            for idx in range(count):
                button = locator.nth(idx)
                try:
                    if button.is_visible(timeout=1000) and button.is_enabled(timeout=1000):
                        button.click(timeout=15000)
                        return True
                except Exception as exc:
                    last_error = str(exc)[:160]
        except Exception as exc:
            last_error = str(exc)[:160]
        page.wait_for_timeout(750)
    raise PlaywrightTimeoutError(
        f"Timed out waiting for RMS {label}. {page_excerpt(page)} last_error={last_error}"
    )

def rms_has_login_inputs(page):
    try:
        if page.locator(RMS_USERNAME_SELECTOR).count() > 0 and page.locator(RMS_PASSWORD_SELECTOR).count() > 0:
            return True
    except Exception:
        pass
    return False

def wait_for_manual_rms_login_or_inputs(page, timeout_ms=300000):
    deadline = time.time() + (timeout_ms / 1000)
    while time.time() < deadline:
        try:
            if not is_rms_login_page(page):
                return "logged_in"
            if rms_has_login_inputs(page):
                return "inputs_ready"
        except Exception:
            pass
        page.wait_for_timeout(1000)
    return "timeout"

def wait_for_rms_login_result(page, timeout_ms=30000):
    deadline = time.time() + (timeout_ms / 1000)
    while time.time() < deadline:
        try:
            if "/login" not in clean(page.url).lower():
                return True
        except Exception:
            pass
        page.wait_for_timeout(500)
    try:
        page.wait_for_load_state("domcontentloaded", timeout=5000)
    except Exception:
        pass
    return "/login" not in clean(page.url).lower()

def login_to_rms(page, login_url, username, password):
    # Required RMS flow step 1.
    # Default: https://rms.reusability.com/login
    # Optional test override: set RMS_START_URL=https://rms.reusability.com
    start_url = clean(os.environ.get("RMS_START_URL")) or normalized_rms_login_url(login_url)
    response = page.goto(start_url, wait_until="domcontentloaded", timeout=90000)
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass
    assert_rms_accessible(page, response, "RMS login page")

    if rms_manual_login_enabled():
        state = wait_for_manual_rms_login_or_inputs(page, timeout_ms=300000)
        if state == "logged_in":
            return
        if state == "timeout":
            diag = write_rms_diagnostic(page, "rms_manual_login_timeout")
            raise RuntimeError(
                "RMS manual login mode timed out. The Edge/Chromium window opened, but EOMS did not detect login completion or visible login fields. "
                "Try refreshing the RMS window or set RMS_BROWSER=edge. Diagnostic: " + clean(str(diag))[:500]
            )

    fill_first_visible(page, RMS_USERNAME_SELECTOR, username, "username")
    fill_first_visible(page, RMS_PASSWORD_SELECTOR, password, "password")
    page.wait_for_timeout(500)
    click_first_visible(
        page,
        'button[type="submit"]:has-text("Sign In"), button:has-text("Sign In"), button:has-text("Login"), input[type="submit"], input[value*="Sign"], button[type="submit"]',
        "login button"
    )
    wait_for_rms_login_result(page, timeout_ms=30000)
    if is_rms_login_page(page):
        # Some RMS builds do not fire the button handler reliably in headless Chromium.
        # Re-fill and press Enter from the password field before treating it as a failed login.
        fill_first_visible(page, RMS_USERNAME_SELECTOR, username, "username", timeout_ms=15000)
        fill_first_visible(page, RMS_PASSWORD_SELECTOR, password, "password", timeout_ms=15000)
        try:
            page.locator(RMS_PASSWORD_SELECTOR).first.press("Enter", timeout=5000)
            wait_for_rms_login_result(page, timeout_ms=30000)
        except Exception:
            try:
                page.keyboard.press("Enter")
                wait_for_rms_login_result(page, timeout_ms=30000)
            except Exception:
                page.wait_for_timeout(3000)
    if is_rms_login_page(page):
        raise RuntimeError(
            "RMS login did not complete. Verify the RMS username/password saved in EOMS Settings. "
            + page_excerpt(page)
        )

def write_rms_diagnostic(page, prefix):
    diag_dir = BASE_DIR / "diagnostics"
    diag_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    screenshot_path = diag_dir / f"{prefix}_{stamp}.png"
    html_path = diag_dir / f"{prefix}_{stamp}.html"
    try:
        page.screenshot(path=str(screenshot_path), full_page=True)
    except Exception:
        screenshot_path = ""
    try:
        html_path.write_text(page.content(), encoding="utf-8", errors="ignore")
    except Exception:
        html_path = ""
    return {"screenshot": str(screenshot_path), "html": str(html_path), "page": page_excerpt(page)}

def collect_bol_links_from_all_pages(page, max_pages=50):
    bol_links = []
    seen = set()

    def harvest_current_page():
        found = []
        try:
            page_snapshot = page.evaluate(
                """() => {
                    const anchors = Array.from(document.querySelectorAll('a')).map((anchor) => {
                        const row = anchor.closest('tr');
                        return {
                            text: (anchor.innerText || anchor.textContent || '').trim(),
                            href: anchor.getAttribute('href') || '',
                            rowText: row ? (row.innerText || row.textContent || '').trim() : ''
                        };
                    });
                    const rows = Array.from(document.querySelectorAll('tr')).map((row) => ({
                        text: (row.innerText || row.textContent || '').trim()
                    }));
                    return {
                        anchors,
                        rows,
                        bodyText: (document.body ? (document.body.innerText || document.body.textContent || '') : '').trim()
                    };
                }"""
            )
        except Exception:
            page_snapshot = {"anchors": [], "rows": [], "bodyText": ""}

        def add_candidate(bol, href="", row_text=""):
            bol = clean(bol)
            if not re.fullmatch(r"\d{4,}", bol) or bol in seen:
                return
            seen.add(bol)
            row_text = clean(row_text)
            due_date = ""
            assigned_date = ""

            # RMS list row usually contains assigned/due dates. Capture first/last date as fallback.
            dates = re.findall(r"\b\d{1,2}/\d{1,2}/\d{4}\b|\b\d{4}-\d{2}-\d{2}\b", row_text)
            if dates:
                assigned_date = normalize_due_date(dates[0])
                due_date = normalize_due_date(dates[-1])

            # If row labels exist, prefer labeled due date.
            labeled_due = find_match(r"Due Date(?:\s*\(PDT\))?\s*[:\s]*([0-9/\-]+)", row_text)
            if labeled_due:
                due_date = normalize_due_date(labeled_due)

            found.append({
                "bol": bol,
                "href": href,
                "row_text": row_text,
                "assigned_date": assigned_date,
                "due_date": due_date
            })

        for anchor in page_snapshot.get("anchors", []):
            try:
                text = clean(anchor.get("text"))
                href = anchor.get("href") or ""
                href_bol = find_match(r"/bills-of-lading/(\d{4,})(?:/print)?", href)
                add_candidate(text if re.fullmatch(r"\d{4,}", text) else href_bol, href, anchor.get("rowText"))
            except Exception:
                pass

        for row in page_snapshot.get("rows", []):
            row_text = clean(row.get("text"))
            for bol in re.findall(r"\b\d{5,8}\b", row_text):
                add_candidate(bol, rms_print_url(bol), row_text)

        # Last fallback: catch visible direct-print links even if RMS renders without table rows.
        body_text = clean(page_snapshot.get("bodyText"))
        for bol in re.findall(r"bills-of-lading/(\d{4,})/print", body_text):
            add_candidate(bol, rms_print_url(bol), body_text[:500])
        return found

    # Try setting rows per page to largest available value first.
    try:
        page.evaluate(
            """() => {
                const selects = Array.from(document.querySelectorAll('select'));
                for (const select of selects) {
                    const labels = Array.from(select.options).map((option) => option.textContent || '');
                    const values = Array.from(select.options).map((option) => option.value || '');
                    const joined = labels.join(' ');
                    if (joined.includes('20') && (joined.includes('50') || joined.includes('100'))) {
                        const targetIndex = labels.findIndex((label) => label.includes('100'));
                        const fallbackIndex = labels.findIndex((label) => label.includes('50'));
                        const index = targetIndex >= 0 ? targetIndex : fallbackIndex;
                        if (index >= 0) {
                            select.value = values[index];
                            select.dispatchEvent(new Event('input', { bubbles: true }));
                            select.dispatchEvent(new Event('change', { bubbles: true }));
                            return true;
                        }
                    }
                }
                return false;
            }"""
        )
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        try:
            page.wait_for_timeout(1500)
        except Exception:
            pass

    for page_index in range(max_pages):
        page.wait_for_timeout(800)
        bol_links.extend(harvest_current_page())

        # Find and click next page arrow. RMS uses pagination with numeric pages and arrows.
        clicked = False

        # First try button/link with text that looks like next arrow.
        possible_next = [
            'button:has-text("›")',
            'button:has-text(">")',
            'a:has-text("›")',
            'a:has-text(">")',
            'button[aria-label*="Next"]',
            'a[aria-label*="Next"]'
        ]

        for selector in possible_next:
            try:
                loc = page.locator(selector).last
                if loc.count() and loc.is_enabled():
                    before = len(bol_links)
                    loc.click()
                    page.wait_for_timeout(1800)
                    clicked = True
                    break
            except Exception:
                pass

        if clicked:
            continue

        # Fallback: click next numeric page if visible.
        try:
            next_number = str(page_index + 2)
            loc = page.get_by_text(next_number, exact=True)
            if loc.count() and loc.first.is_visible():
                loc.first.click()
                page.wait_for_timeout(1800)
                clicked = True
        except Exception:
            pass

        if not clicked:
            break

    # De-dupe just in case.
    unique = []
    seen2 = set()
    for item in bol_links:
        if item["bol"] not in seen2:
            unique.append(item)
            seen2.add(item["bol"])
    return unique

def import_printable_bol_item(item, existing_keys=None, existing_bols=None):
    stores = read_json(STORES_FILE)
    replaced = False
    now = datetime.now().isoformat(timespec="seconds")
    item["last_seen_in_rms_at"] = now
    item["rms_status"] = "Open in RMS"
    item["rms_missing_since"] = ""
    for idx, existing in enumerate(stores):
        if clean(existing.get("bol")) == clean(item.get("bol")):
            item["id"] = existing.get("id", item.get("id"))
            for key in [
                "assigned_driver", "assigned_driver_phone", "assigned_at",
                "collected_racks", "collected_pieces", "variance",
                "variance_review", "pieces_variance", "completed_at",
                "closeout_updated_at", "closeout_updated_by"
            ]:
                if existing.get(key) not in (None, "") and item.get(key) in (None, ""):
                    item[key] = existing.get(key)
            if existing.get("status") in {"Assigned", "Dispatched", "Completed"}:
                item["status"] = existing.get("status")
            stores[idx] = item
            replaced = True
            break
    if not replaced:
        stores.append(item)
    write_json(STORES_FILE, stores)
    if existing_keys is not None and existing_bols is not None:
        track_bol_key(item, existing_keys, existing_bols)
    update_queue_from_item(item)
    return "updated" if replaced else "imported"

def mark_stores_missing_from_rms(open_bols):
    open_bols = {clean(b) for b in open_bols if clean(b)}
    stores = read_json(STORES_FILE)
    now = datetime.now().isoformat(timespec="seconds")
    marked = 0
    both_closed = 0

    for store in stores:
        bol = clean(store.get("bol"))
        if not bol or bol in open_bols:
            if bol in open_bols:
                store["last_seen_in_rms_at"] = now
                store["rms_status"] = "Open in RMS"
                store["rms_missing_since"] = ""
            continue

        if not (store.get("rms_url") or store.get("pdf_path") or store.get("printable_path")):
            continue

        if clean(store.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}:
            continue

        store["rms_status"] = "Closed in RMS" if store.get("status") == "Completed" else "Missing from RMS"
        store["rms_missing_since"] = store.get("rms_missing_since") or now
        store["closed_source"] = "Both" if store.get("status") == "Completed" else "RMS"
        store["updated_at"] = now
        if store.get("pdf_path"):
            store["pdf_path"] = move_pdf(store["pdf_path"], "RMS_Closed")
        marked += 1
        if store.get("closed_source") == "Both":
            both_closed += 1

    if marked:
        write_json(STORES_FILE, stores)
    return {"rms_missing": marked, "both_closed": both_closed}

def open_printable_and_extract(page, bol_link):
    bol = bol_link.get("bol")
    href = bol_link.get("href") or ""

    # Navigate to BOL detail.
    if href.startswith("http"):
        page.goto(href, wait_until="networkidle", timeout=60000)
    elif href.startswith("/"):
        page.goto("https://rms.reusability.com" + href, wait_until="networkidle", timeout=60000)
    else:
        # Fallback click by BOL text from list page if relative URL missing.
        page.get_by_text(bol, exact=True).click()
        page.wait_for_load_state("networkidle", timeout=60000)

    # Click printable BOL link.
    printable_clicked = False
    try:
        page.get_by_text(re.compile("View Printable Bill of Lading|Print this Bill of Lading", re.I)).first.click()
        page.wait_for_load_state("networkidle", timeout=60000)
        printable_clicked = True
    except Exception:
        pass

    if not printable_clicked:
        try:
            page.locator('a:has-text("Printable")').first.click()
            page.wait_for_load_state("networkidle", timeout=60000)
            printable_clicked = True
        except Exception:
            pass

    text = page.locator("body").inner_text(timeout=60000)
    html = page.content()
    item = extract_printable_bol_from_text(text, page.url)
    if not item.get("bol"):
        item["bol"] = bol
        item.setdefault("review_warnings", []).append("BOL fallback from RMS list")
        item["review_reasons"] = essential_review_reasons(item)
        item["status"] = "Need Review" if item["review_reasons"] else "Unassigned"
    item["pdf_path"] = save_printable_pdf(page, item, html)
    return item


def scan_rms_queue_with_playwright(headless=True):
    settings_data = read_json(SETTINGS_FILE)

    username = clean(settings_data.get("rms_username"))
    password = clean(settings_data.get("rms_password"))
    login_url = normalized_rms_login_url(settings_data.get("rms_login_url"))
    bol_url = normalized_rms_bol_list_url(settings_data.get("rms_bol_url"))

    if not username or not password:
        return {
            "ok": False,
            "status": "MISSING CREDENTIALS",
            "message": "Save RMS username and password in Settings first.",
            "found": 0,
            "new": 0,
            "existing": 0
        }

    existing_stores = read_json(STORES_FILE)
    existing_bols = {clean(s.get("bol")) for s in existing_stores if clean(s.get("bol"))}
    current_queue = read_json(RMS_QUEUE_FILE)
    queue_by_bol = {clean(q.get("bol")): q for q in current_queue if clean(q.get("bol"))}

    with sync_playwright() as p:
        browser, context = launch_rms_browser_context(p, headless=headless)
        page = context.new_page()

        try:
            login_to_rms(page, login_url, username, password)

            # Required RMS flow step 2:
            #   https://rms.reusability.com/bills-of-lading
            bol_url = normalized_rms_bol_list_url(bol_url)
            response = page.goto(bol_url, wait_until="networkidle", timeout=60000)
            assert_rms_accessible(page, response, "RMS BOL list")
            if is_rms_login_page(page):
                raise RuntimeError(
                    "RMS redirected back to login when opening the BOL list. "
                    "Verify RMS credentials and that the account can access Bills of Lading. "
                    + page_excerpt(page)
                )
            bol_links = collect_bol_links_from_all_pages(page)

            new_count = 0
            existing_count = 0

            for item in bol_links:
                bol = clean(item.get("bol"))
                if not bol:
                    continue

                status = "Imported" if bol in existing_bols else "New"
                if bol in existing_bols:
                    existing_count += 1
                else:
                    new_count += 1

                if bol in queue_by_bol:
                    # preserve ignored/manual status unless already imported
                    old = queue_by_bol[bol]
                    if old.get("queue_status") in ["Ignored", "Need Review"] and status != "Imported":
                        status = old.get("queue_status")
                    old.update({
                        "href": item.get("href", ""),
                        "queue_status": status,
                        "due_date": item.get("due_date", old.get("due_date", "")),
                        "assigned_date": item.get("assigned_date", old.get("assigned_date", "")),
                        "row_text": item.get("row_text", old.get("row_text", "")),
                        "last_seen": datetime.now().isoformat(timespec="seconds")
                    })
                else:
                    queue_by_bol[bol] = {
                        "id": str(uuid4()),
                        "bol": bol,
                        "href": item.get("href", ""),
                        "queue_status": status,
                        "store_name": "",
                        "origin": "",
                        "city": "",
                        "state": "",
                        "hub": "",
                        "expected_racks": "",
                        "weight": "",
                        "due_date": item.get("due_date", ""),
                        "assigned_date": item.get("assigned_date", ""),
                        "row_text": item.get("row_text", ""),
                        "last_seen": datetime.now().isoformat(timespec="seconds"),
                        "created_at": datetime.now().isoformat(timespec="seconds")
                    }

            queue = list(queue_by_bol.values())
            queue.sort(key=lambda x: x.get("bol", ""), reverse=True)
            write_json(RMS_QUEUE_FILE, queue)

            diag_dir = BASE_DIR / "diagnostics"
            diag_dir.mkdir(exist_ok=True)
            page.screenshot(path=str(diag_dir / f"rms_queue_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"), full_page=True)

            rms_debug_hold(page)
            close_rms_browser(browser, context)

            return {
                "ok": True,
                "status": "QUEUE UPDATED",
                "message": f"RMS queue updated. Found {len(bol_links)} BOLs. New {new_count}. Already imported {existing_count}.",
                "found": len(bol_links),
                "new": new_count,
                "existing": existing_count
            }

        except Exception as e:
            close_rms_browser(browser, context)
            return {
                "ok": False,
                "status": "ERROR",
                "message": f"RMS queue scan error: {str(e)[:300]}",
                "found": 0,
                "new": 0,
                "existing": 0
            }


def upsert_store_by_bol(item):
    stores = read_json(STORES_FILE)
    replaced = False
    now = datetime.now().isoformat(timespec="seconds")
    item["last_seen_in_rms_at"] = now
    item["rms_status"] = "Open in RMS"
    item["rms_missing_since"] = ""
    for idx, existing in enumerate(stores):
        if clean(existing.get("bol")) == clean(item.get("bol")):
            item["id"] = existing.get("id", item.get("id"))
            if existing.get("status") in {"Assigned", "Dispatched", "Completed"}:
                item["status"] = existing.get("status")
            stores[idx] = item
            replaced = True
            break
    if not replaced:
        stores.append(item)
    write_json(STORES_FILE, stores)
    return item

def update_queue_from_item(item):
    queue = read_json(RMS_QUEUE_FILE)
    for q in queue:
        if clean(q.get("bol")) == clean(item.get("bol")):
            q["queue_status"] = "Imported" if item.get("status") == "Unassigned" else "Need Review"
            q["store_name"] = item.get("store_name", "")
            q["origin"] = item.get("origin", "")
            q["city"] = item.get("city", "")
            q["state"] = item.get("state", "")
            q["address"] = item.get("address", "")
            q["contact"] = item.get("contact", "")
            q["origin_name"] = item.get("origin_name", "")
            q["origin_address"] = item.get("origin_address", "")
            q["origin_contact"] = item.get("origin_contact", "")
            q["hub"] = item.get("hub", "")
            q["expected_racks"] = item.get("expected_racks", "")
            q["weight"] = item.get("weight", "")
            q["due_date"] = item.get("due_date", "")
            q["assigned_date"] = item.get("assigned_date", "")
            q["pdf_path"] = item.get("pdf_path", "")
            q["geocode_status"] = item.get("geocode_status", "")
            q["full_address"] = item.get("full_address", "")
            q["imported_at"] = datetime.now().isoformat(timespec="seconds")
            break
    write_json(RMS_QUEUE_FILE, queue)

def import_selected_queue_bols(bol_numbers, headless=True):
    settings_data = read_json(SETTINGS_FILE)

    username = clean(settings_data.get("rms_username"))
    password = clean(settings_data.get("rms_password"))
    login_url = normalized_rms_login_url(settings_data.get("rms_login_url"))

    if not username or not password:
        return {
            "ok": False,
            "status": "MISSING CREDENTIALS",
            "message": "Save RMS username and password in Settings first.",
            "imported": 0,
            "skipped": 0,
            "need_review": 0,
            "errors": []
        }

    queue = read_json(RMS_QUEUE_FILE)
    queue_by_bol = {clean(q.get("bol")): q for q in queue}

    imported = 0
    need_review = 0
    errors = []

    with sync_playwright() as p:
        browser, context = launch_rms_browser_context(p, headless=headless)
        page = context.new_page()

        try:
            login_to_rms(page, login_url, username, password)

            for bol in bol_numbers:
                bol = clean(bol)
                if not bol:
                    continue

                try:
                    direct_print_url = rms_print_url(bol)
                    response = page.goto(direct_print_url, wait_until="networkidle", timeout=60000)
                    assert_rms_accessible(page, response, f"RMS printable BOL {bol}")

                    body_text = page.locator("body").inner_text(timeout=60000)
                    html = page.content()

                    item = extract_printable_bol_from_text(body_text, page.url)

                    if not item.get("bol"):
                        item["bol"] = bol
                    if clean(item.get("bol")) != bol:
                        item["bol"] = bol

                    queue_item = queue_by_bol.get(bol, {})
                    if not item.get("due_date") and queue_item.get("due_date"):
                        item["due_date"] = queue_item.get("due_date")
                    if not item.get("assigned_date") and queue_item.get("assigned_date"):
                        item["assigned_date"] = queue_item.get("assigned_date")

                    item["pdf_path"] = save_printable_pdf(page, item, html)

                    # Force automatic geocoding during every Import/Re-import.
                    # This replaces city-center fallback coordinates when an Azure Maps key is saved.
                    lat, lng, geocode_status, full_address = resolve_store_coordinates(
                        item.get("address") or item.get("origin_address") or "",
                        item.get("city") or item.get("origin_city") or "",
                        item.get("state") or item.get("origin_state") or "TX",
                        item.get("zip") or item.get("origin_zip") or ""
                    )
                    item["lat"] = lat
                    item["lng"] = lng
                    item["full_address"] = full_address
                    item["geocode_status"] = geocode_status

                    upsert_store_by_bol(item)
                    update_queue_from_item(item)

                    imported += 1
                    if item.get("status") == "Need Review":
                        need_review += 1

                except Exception as e:
                    errors.append(f"BOL {bol}: {str(e)[:180]}")
                    queue = read_json(RMS_QUEUE_FILE)
                    for q in queue:
                        if clean(q.get("bol")) == bol:
                            q["queue_status"] = "Need Review"
                            q["updated_at"] = datetime.now().isoformat(timespec="seconds")
                            break
                    write_json(RMS_QUEUE_FILE, queue)

            close_rms_browser(browser, context)

            return {
                "ok": True,
                "status": "IMPORT SELECTED COMPLETE",
                "message": f"Import/Re-import complete. Processed {imported}. Need Review {need_review}.",
                "imported": imported,
                "skipped": 0,
                "need_review": need_review,
                "errors": errors[:10]
            }

        except Exception as e:
            try:
                if not rms_manual_login_enabled():
                    context.close()
            except Exception:
                pass
            try:
                close_rms_browser(browser, context)
            except Exception:
                pass

            return {
                "ok": False,
                "status": "ERROR",
                "message": f"Import/Re-import selected error: {str(e)[:300]}",
                "imported": imported,
                "skipped": 0,
                "need_review": need_review,
                "errors": errors[:10]
            }

def rms_full_import_with_playwright(headless=True, max_bols=0):
    settings_data = read_json(SETTINGS_FILE)

    username = clean(settings_data.get("rms_username"))
    password = clean(settings_data.get("rms_password"))
    login_url = normalized_rms_login_url(settings_data.get("rms_login_url"))
    bol_url = normalized_rms_bol_list_url(settings_data.get("rms_bol_url"))

    if not rms_manual_login_enabled() and (not username or not password):
        return {
            "ok": False,
            "status": "MISSING CREDENTIALS",
            "message": "Save RMS username and password in Settings first.",
            "imported": 0,
            "skipped": 0,
            "need_review": 0,
            "bol_count": 0
        }

    existing = read_json(STORES_FILE)
    existing_keys = {bol_duplicate_key(s) for s in existing if clean(s.get("bol")) or clean(s.get("origin"))}
    existing_bols = {clean(s.get("bol")) for s in existing if clean(s.get("bol"))}

    imported = 0
    updated = 0
    skipped = 0
    need_review = 0
    errors = []

    with sync_playwright() as p:
        browser, context = launch_rms_browser_context(p, headless=headless)
        page = context.new_page()
        try:
            bol_url = normalized_rms_bol_list_url(bol_url)
            if rms_manual_login_enabled():
                open_rms_bol_list_with_manual_login(page, bol_url)
            else:
                login_to_rms(page, login_url, username, password)

                # Required RMS flow step 2:
                #   https://rms.reusability.com/bills-of-lading
                response = page.goto(bol_url, wait_until="networkidle", timeout=60000)
                assert_rms_accessible(page, response, "RMS BOL list")
                if is_rms_login_page(page):
                    raise RuntimeError(
                        "RMS redirected back to login when opening the BOL list. "
                        "Verify RMS credentials and that the account can access Bills of Lading. "
                        + page_excerpt(page)
                    )
            bol_links = collect_bol_links_from_all_pages(page)

            diagnostic = None
            if not bol_links:
                diagnostic = write_rms_diagnostic(page, "rms_no_bol_links")

            if max_bols and max_bols > 0:
                bol_links = bol_links[:max_bols]
            open_bols = {clean(link.get("bol")) for link in bol_links if clean(link.get("bol"))}

            # Diagnostic screenshot after pagination scan.
            diag_dir = BASE_DIR / "diagnostics"
            diag_dir.mkdir(exist_ok=True)
            page.screenshot(path=str(diag_dir / f"rms_pagination_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"), full_page=True)

            for link in bol_links:
                try:
                    # Reuse the same visible RMS tab for each BOL so headed
                    # local imports do not open a new window for every stop.
                    detail_page = page
                    # Required RMS flow step 3 for each BOL, example:
                    #   https://rms.reusability.com/bills-of-lading/951807/print
                    direct_print_url = rms_print_url(link["bol"])
                    response = detail_page.goto(direct_print_url, wait_until="networkidle", timeout=60000)
                    assert_rms_accessible(detail_page, response, f"RMS printable BOL {link['bol']}")
                    if is_rms_login_page(detail_page):
                        raise RuntimeError("RMS redirected printable BOL back to login. " + page_excerpt(detail_page))

                    text = detail_page.locator("body").inner_text(timeout=60000)
                    html = detail_page.content()
                    item = extract_printable_bol_from_text(text, detail_page.url)
                    if not item.get("bol"):
                        item["bol"] = link["bol"]
                        item.setdefault("review_warnings", []).append("BOL fallback from RMS list")
                        item["review_reasons"] = essential_review_reasons(item)
                        item["status"] = "Need Review" if item["review_reasons"] else "Unassigned"

                    item["pdf_path"] = save_printable_pdf(detail_page, item, html)
                    action = import_printable_bol_item(item, existing_keys, existing_bols)
                    if action == "updated":
                        updated += 1
                    else:
                        imported += 1
                    if item.get("status") == "Need Review":
                        need_review += 1

                except Exception as e:
                    errors.append(f"{link.get('bol')}: {str(e)[:120]}")

            rms_closeout = mark_stores_missing_from_rms(open_bols) if bol_links else {"rms_missing": 0, "both_closed": 0}
            rms_debug_hold(page)
            close_rms_browser(browser, context)

            message = f"RMS import complete using correct RMS flow: login → bills-of-lading list → individual /print pages. Scanned {len(bol_links)} BOLs. Imported {imported}. Updated {updated}. Skipped {skipped}. Need Review {need_review}. RMS missing/closed {rms_closeout['rms_missing']}."
            if diagnostic:
                message = "RMS login/page load completed, but no BOL links were detected. Check RMS credentials, BOL URL, filters, and diagnostics. " + diagnostic.get("page", "")
            all_failed = bool(bol_links) and not imported and not updated and bool(errors)
            if all_failed:
                message = f"RMS import scanned {len(bol_links)} BOLs but could not download any printable BOLs. First error: {errors[0]}"

            return {
                "ok": not all_failed,
                "status": "IMPORT COMPLETE",
                "message": message,
                "bol_count": len(bol_links),
                "found": len(bol_links),
                "imported": imported,
                "updated": updated,
                "skipped": skipped,
                "need_review": need_review,
                "rms_missing": rms_closeout["rms_missing"],
                "both_closed": rms_closeout["both_closed"],
                "failed": len(errors),
                "diagnostic": diagnostic,
                "errors": errors[:10]
            }

        except Exception as e:
            rms_debug_hold(page)
            close_rms_browser(browser, context)
            return {
                "ok": False,
                "status": "ERROR",
                "message": f"RMS full import error: {str(e)[:300]}",
                "imported": imported,
                "updated": updated,
                "skipped": skipped,
                "need_review": need_review,
                "bol_count": 0,
                "found": 0,
                "failed": max(1, len(errors)),
                "errors": errors[:10]
            }




@app.route("/bol-live/<bol_number>")
def bol_live_printable(bol_number):
    saved_path = find_saved_bol_path(bol_number)
    if saved_path:
        saved_record = next((s for s in read_json(STORES_FILE) if clean(s.get("bol")) == clean(bol_number)), {"bol": bol_number})
        return render_saved_bol(saved_record, saved_path, auto_print=False)

    settings_data = read_json(SETTINGS_FILE)

    username = clean(settings_data.get("rms_username"))
    password = clean(settings_data.get("rms_password"))
    login_url = normalized_rms_login_url(settings_data.get("rms_login_url"))

    if not rms_manual_login_enabled() and (not username or not password):
        return render_template(
            "bol_not_found.html",
            bol=bol_number,
            path="RMS username/password missing. Save RMS credentials in Settings first."
        )

    direct_print_url = rms_print_url(bol_number)

    with sync_playwright() as p:
        browser, context = launch_rms_browser_context(p, headless=rms_headless(False))
        page = context.new_page()

        try:
            if rms_manual_login_enabled():
                open_rms_bol_list_with_manual_login(page, normalized_rms_bol_list_url(settings_data.get("rms_bol_url")))
            else:
                login_to_rms(page, login_url, username, password)

            # Go straight to the RMS printable URL.
            response = page.goto(direct_print_url, wait_until="networkidle", timeout=60000)
            assert_rms_accessible(page, response, f"RMS printable BOL {bol_number}")

            body_text, html = wait_for_printable_bol_content(page, bol_number)
            if is_rms_login_page(page):
                raise RuntimeError("RMS redirected Live Print back to login. " + page_excerpt(page))

            if is_rms_wait_page(body_text, html):
                saved_path = find_saved_bol_path(bol_number)
                if saved_path:
                    saved_record = next((s for s in read_json(STORES_FILE) if clean(s.get("bol")) == clean(bol_number)), {"bol": bol_number})
                    try:
                        if not rms_manual_login_enabled():
                            context.close()
                    except Exception:
                        pass
                    close_rms_browser(browser, context)
                    return render_saved_bol(saved_record, saved_path, auto_print=False)
                raise RuntimeError("RMS only returned its 'Please wait' print shell and no saved copy was found.")

            # Save backup snapshot.
            backup_dir = month_folder("RMS_Backup")
            backup_file = backup_dir / f"DIRECT_PRINTABLE_BOL_{safe_part(bol_number)}_{datetime.now().strftime('%H%M%S')}.html"
            backup_file.write_text(html, encoding="utf-8", errors="ignore")

            rms_debug_hold(page)
            close_rms_browser(browser, context)

            saved_record = next((s for s in read_json(STORES_FILE) if clean(s.get("bol")) == clean(bol_number)), {"bol": bol_number})
            return f"{bol_print_styles()}{bol_edit_summary_html(saved_record)}{html}"

        except Exception as e:
            try:
                if not rms_manual_login_enabled():
                    context.close()
            except Exception:
                pass
            try:
                close_rms_browser(browser, context)
            except Exception:
                pass

            saved_path = find_saved_bol_path(bol_number)
            if saved_path:
                saved_record = next((s for s in read_json(STORES_FILE) if clean(s.get("bol")) == clean(bol_number)), {"bol": bol_number})
                return render_saved_bol(saved_record, saved_path, auto_print=False)

            return render_template(
                "bol_not_found.html",
                bol=bol_number,
                path=f"Direct printable RMS error: {str(e)[:300]}"
            )


@app.route("/bol-print/<store_id>")
def bol_print(store_id):
    stores = read_json(STORES_FILE)
    queue = read_json(RMS_QUEUE_FILE)

    found = None
    for s in stores:
        if s.get("id") == store_id or clean(s.get("bol")) == store_id:
            found = s
            break

    if not found:
        for q in queue:
            if q.get("id") == store_id or clean(q.get("bol")) == store_id:
                found = q
                break

    if not found:
        abort(404)

    path = found.get("pdf_path") or found.get("printable_path") or ""
    if not path:
        bol = safe_part(found.get("bol"))
        matches = list(BOL_DIR.rglob(f"*{bol}*.html")) + list(BOL_DIR.rglob(f"*{bol}*.pdf"))
        if matches:
            path = str(matches[0])

    if not path or not Path(path).exists():
        return render_template("bol_not_found.html", bol=found.get("bol", ""), path=path)

    return render_saved_bol(found, path, auto_print=True)

@app.route("/bol-view/<store_id>")
def bol_view(store_id):
    stores = read_json(STORES_FILE)
    queue = read_json(RMS_QUEUE_FILE)

    found = None
    for s in stores:
        if s.get("id") == store_id or clean(s.get("bol")) == store_id:
            found = s
            break

    if not found:
        for q in queue:
            if q.get("id") == store_id or clean(q.get("bol")) == store_id:
                found = q
                break

    if not found:
        abort(404)

    path = found.get("pdf_path") or found.get("printable_path") or ""
    if not path:
        # Try to find saved file by BOL number
        bol = safe_part(found.get("bol"))
        matches = list(BOL_DIR.rglob(f"*{bol}*.html")) + list(BOL_DIR.rglob(f"*{bol}*.pdf"))
        if matches:
            path = str(matches[0])

    if not path or not Path(path).exists():
        return render_template("bol_not_found.html", bol=found.get("bol", ""), path=path)

    return render_saved_bol(found, path, auto_print=False)

@app.route("/api/approve-review-force", methods=["POST"])
def api_approve_review_force():
    data = request.get_json(force=True)
    store_id = data.get("store_id")
    hub = data.get("hub")

    stores = read_json(STORES_FILE)
    updated = None

    for store in stores:
        if store.get("id") == store_id:
            store["hub"] = hub or store.get("hub") or "San Antonio"
            store["status"] = "Unassigned"
            store["review_reasons"] = []
            if not store.get("lat") or not store.get("lng"):
                store["lat"], store["lng"] = HUBS.get(store["hub"], HUBS["San Antonio"])["lat"], HUBS.get(store["hub"], HUBS["San Antonio"])["lng"]
            updated = store
            break

    write_json(STORES_FILE, stores)
    audit("Force Approve Need Review", {"store_id": store_id, "hub": hub})

    return jsonify({"ok": True, "store": updated})


@app.route("/api/rms/repair-bol/<bol_number>", methods=["POST"])
def api_rms_repair_bol(bol_number):
    settings_data = read_json(SETTINGS_FILE)
    username = clean(settings_data.get("rms_username"))
    password = clean(settings_data.get("rms_password"))
    login_url = normalized_rms_login_url(settings_data.get("rms_login_url"))

    if not username or not password:
        return jsonify({"ok": False, "message": "Save RMS credentials first."})

    with sync_playwright() as p:
        browser = launch_chromium_with_repair(p, headless=True)
        context = new_rms_context(browser)
        page = context.new_page()
        try:
            login_to_rms(page, login_url, username, password)

            direct_print_url = rms_print_url(bol_number)
            response = page.goto(direct_print_url, wait_until="networkidle", timeout=60000)
            assert_rms_accessible(page, response, f"RMS printable BOL {bol_number}")

            body_text = page.locator("body").inner_text(timeout=60000)
            html = page.content()
            item = extract_printable_bol_from_text(body_text, page.url)
            item["pdf_path"] = save_printable_pdf(page, item, html)

            stores = read_json(STORES_FILE)
            replaced = False
            for idx, s in enumerate(stores):
                if clean(s.get("bol")) == clean(bol_number):
                    item["id"] = s.get("id", item["id"])
                    stores[idx] = item
                    replaced = True
                    break
            if not replaced:
                stores.append(item)

            write_json(STORES_FILE, stores)

            update_queue_from_item(item)

            close_rms_browser(browser, context)
            return jsonify({"ok": True, "item": item, "message": f"BOL {bol_number} repaired. Status: {item.get('status')}. Racks: {item.get('expected_racks')}"})
        except Exception as e:
            try:
                if not rms_manual_login_enabled():
                    context.close()
            except Exception:
                pass
            try:
                close_rms_browser(browser, context)
            except Exception:
                pass
            return jsonify({"ok": False, "message": str(e)[:300]})

@app.route("/rms-queue")
def rms_queue():
    queue = filter_queue_for_user(read_json(RMS_QUEUE_FILE))
    stores = filter_stores_for_user(read_json(STORES_FILE))
    imported_bols = {clean(s.get("bol")) for s in stores if clean(s.get("bol"))}

    for q in queue:
        if clean(q.get("bol")) in imported_bols and q.get("queue_status") != "Ignored":
            q["queue_status"] = "Imported"

    write_json(RMS_QUEUE_FILE, queue)
    return render_template("rms_queue.html", queue=queue)

@app.route("/api/rms/queue-refresh", methods=["POST"])
def api_rms_queue_refresh():
    result = scan_rms_queue_with_playwright(headless=rms_headless(True))

    history = read_json(SYNC_HISTORY_FILE)
    result.update({
        "id": str(uuid4()),
        "time": datetime.now().isoformat(timespec="seconds"),
        "action": "Refresh RMS Queue"
    })
    history.append(result)
    write_json(SYNC_HISTORY_FILE, history)
    audit("Refresh RMS Queue", result)

    return jsonify({"ok": result.get("ok", False), "result": result})

@app.route("/api/rms/queue-import-selected", methods=["POST"])
def api_rms_queue_import_selected():
    payload = request.get_json(force=True)
    bol_numbers = [clean(x) for x in payload.get("bols", []) if clean(x)]

    if not bol_numbers:
        return jsonify({"ok": False, "result": {"message": "Select at least one BOL."}})

    result = import_selected_queue_bols(bol_numbers, headless=rms_headless(True))

    history = read_json(SYNC_HISTORY_FILE)
    result.update({
        "id": str(uuid4()),
        "time": datetime.now().isoformat(timespec="seconds"),
        "action": "Import Selected RMS Queue"
    })
    history.append(result)
    write_json(SYNC_HISTORY_FILE, history)
    audit("Import Selected RMS Queue", result)

    return jsonify({"ok": result.get("ok", False), "result": result})

@app.route("/api/rms/queue-update-status", methods=["POST"])
def api_rms_queue_update_status():
    payload = request.get_json(force=True)
    bol_numbers = [clean(x) for x in payload.get("bols", []) if clean(x)]
    status = clean(payload.get("status")) or "New"

    queue = read_json(RMS_QUEUE_FILE)
    updated = 0
    for q in queue:
        if clean(q.get("bol")) in bol_numbers:
            q["queue_status"] = status
            q["updated_at"] = datetime.now().isoformat(timespec="seconds")
            updated += 1

    write_json(RMS_QUEUE_FILE, queue)
    audit("Update RMS Queue Status", {"status": status, "updated": updated})

    return jsonify({"ok": True, "updated": updated})

@app.route("/rms-sync")
@admin_required
def rms_sync():
    settings_data = read_json(SETTINGS_FILE)
    history = read_json(SYNC_HISTORY_FILE)
    return render_template("rms_sync.html", settings=settings_data, history=history)

@app.route("/api/rms/test-connection", methods=["POST"])
def api_rms_test_connection():
    history = read_json(SYNC_HISTORY_FILE)

    result = rms_login_with_playwright(headless=rms_headless(True))
    result.update({
        "id": str(uuid4()),
        "time": datetime.now().isoformat(timespec="seconds"),
        "action": "Test RMS Login"
    })

    settings_data = read_json(SETTINGS_FILE)
    settings_data["rms_connection_status"] = result.get("status", "Unknown")
    if result.get("ok"):
        settings_data["last_rms_login"] = datetime.now().isoformat(timespec="seconds")
    write_json(SETTINGS_FILE, settings_data)

    history.append(result)
    write_json(SYNC_HISTORY_FILE, history)
    audit("RMS Login Test", result)

    return jsonify({"ok": result.get("ok", False), "result": result})

@app.route("/api/rms/sync-open-bols", methods=["POST"])
def api_rms_sync_open_bols():
    history = read_json(SYNC_HISTORY_FILE)

    # Safety default: import all found BOLs. Set max_bols in JSON for testing if needed.
    payload = request.get_json(silent=True) or {}
    max_bols = int(payload.get("max_bols") or 0)

    try:
        result = rms_full_import_with_playwright(headless=rms_headless(True), max_bols=max_bols)
    except RMSAccessError as exc:
        result = rms_blocked_result(str(exc), action="Full RMS Multi-Page Import")
    except Exception as exc:
        result = {
            "ok": False,
            "status": "AUTO GRAB ERROR",
            "message": f"Auto Grab failed before RMS import completed: {str(exc)[:500]}",
            "found": 0,
            "imported": 0,
            "updated": 0,
            "failed": 1,
        }
    result.update({
        "id": str(uuid4()),
        "time": datetime.now().isoformat(timespec="seconds"),
        "action": "Full RMS Multi-Page Import"
    })

    history.append(result)
    write_json(SYNC_HISTORY_FILE, history)
    audit("Full RMS Multi-Page Import", result)

    return jsonify({"ok": result.get("ok", False), "result": result})


@app.route("/api/rms/auto-grab-bols", methods=["POST"])
def api_rms_auto_grab_bols():
    """Dashboard one-click RMS grabber.
    Uses the same full RMS multi-page import engine as RMS Sync, but gives the
    dashboard button a dedicated endpoint and simpler response target.
    """
    try:
        history = read_json(SYNC_HISTORY_FILE)
        payload = request.get_json(silent=True) or {}
        max_bols = int(payload.get("max_bols") or 0)

        try:
            result = rms_full_import_with_playwright(headless=rms_headless(True), max_bols=max_bols)
        except RMSAccessError as exc:
            result = rms_blocked_result(str(exc), action="Dashboard Auto Grab BOLs")
        except Exception as exc:
            result = {
                "ok": False,
                "status": "AUTO GRAB ERROR",
                "message": f"RMS Auto Grab failed before import completed: {str(exc)[:700]}",
                "found": 0,
                "imported": 0,
                "updated": 0,
                "failed": 1,
            }
        result.update({
            "id": str(uuid4()),
            "time": datetime.now().isoformat(timespec="seconds"),
            "action": "Dashboard Auto Grab BOLs"
        })

        history.append(result)
        write_json(SYNC_HISTORY_FILE, history)
        audit("Dashboard Auto Grab BOLs", result)

        return jsonify({"ok": result.get("ok", False), "result": result})
    except Exception as exc:
        return jsonify({
            "ok": False,
            "result": {
                "ok": False,
                "status": "AUTO GRAB SERVER ERROR",
                "message": f"Auto Grab server error: {str(exc)[:700]}",
                "found": 0,
                "imported": 0,
                "updated": 0,
                "failed": 1,
            }
        }), 500

@app.route("/archive")
@dispatch_required
def archive():
    stores = filter_stores_for_user(read_json(STORES_FILE))
    routes = filter_routes_for_user(read_json(ROUTES_FILE))

    q = clean(request.args.get("q")).lower()
    city = clean(request.args.get("city"))
    driver = clean(request.args.get("driver"))
    date_from = clean(request.args.get("from"))
    date_to = clean(request.args.get("to"))
    archive_type = clean(request.args.get("type")) or "all"

    archived_stores = [
        s for s in stores
        if (s.get("status") or "") == "Completed"
        or clean(s.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}
    ]
    completed_routes = [r for r in routes if (r.get("status") or "") == "Completed"]

    def date_in_range(value):
        value = clean(value)[:10]
        if date_from and value and value < date_from:
            return False
        if date_to and value and value > date_to:
            return False
        return True

    def store_matches(store):
        rms_status = clean(store.get("rms_status"))
        eoms_completed = (store.get("status") or "") == "Completed"
        if archive_type == "eoms_completed" and not eoms_completed:
            return False
        if archive_type == "rms_missing" and rms_status not in {"Missing from RMS", "Closed in RMS"}:
            return False
        if archive_type == "both" and not (eoms_completed and rms_status in {"Missing from RMS", "Closed in RMS"}):
            return False
        haystack = " ".join([
            clean(store.get("bol")), clean(store.get("origin")),
            clean(store.get("store_name")), clean(store.get("city")),
            clean(store.get("assigned_driver")), clean(store.get("due_date")),
            clean(store.get("completed_at")), rms_status,
        ]).lower()
        if q and q not in haystack:
            return False
        if city and city != clean(store.get("city")) and city != clean(store.get("hub")):
            return False
        if driver and driver.lower() not in clean(store.get("assigned_driver")).lower():
            return False
        archive_date = store.get("completed_at") or store.get("rms_missing_since") or store.get("updated_at")
        if not date_in_range(archive_date):
            return False
        return True

    def route_matches(route):
        haystack = " ".join([
            clean(route.get("route_number")), clean(route.get("driver")),
            clean(route.get("hub")), clean(route.get("completed_at")),
        ]).lower()
        if q and q not in haystack:
            return False
        if driver and driver.lower() not in clean(route.get("driver")).lower():
            return False
        if city and city != clean(route.get("hub")):
            return False
        if not date_in_range(route.get("completed_at")):
            return False
        return True

    completed_stores = [s for s in archived_stores if store_matches(s)]
    completed_routes = [r for r in completed_routes if route_matches(r)]

    completed_stores.sort(key=lambda s: clean(s.get("completed_at")) or clean(s.get("rms_missing_since")) or clean(s.get("created_at")), reverse=True)
    completed_routes.sort(key=lambda r: clean(r.get("completed_at")) or clean(r.get("created_at")), reverse=True)

    city_options = sorted({
        clean(s.get("city")) for s in stores if clean(s.get("city"))
    } | {
        clean(s.get("hub")) for s in stores if clean(s.get("hub"))
    })
    driver_options = sorted({
        clean(s.get("assigned_driver")) for s in stores if clean(s.get("assigned_driver"))
    } | {
        clean(r.get("driver")) for r in routes if clean(r.get("driver"))
    })

    totals = {
        "stores": len(completed_stores),
        "routes": len(completed_routes),
        "eoms_completed": len([s for s in completed_stores if (s.get("status") or "") == "Completed"]),
        "rms_missing": len([s for s in completed_stores if clean(s.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}]),
        "both_closed": len([s for s in completed_stores if (s.get("status") or "") == "Completed" and clean(s.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}]),
        "expected_racks": round(sum(num(s.get("expected_racks")) for s in completed_stores), 2),
        "collected_racks": round(sum(num(s.get("collected_racks")) for s in completed_stores), 2),
        "variance": round(sum(num(s.get("variance")) for s in completed_stores), 2),
        "expected_pieces": round(sum(num(s.get("expected_pieces")) or (num(s.get("expected_racks")) * PIECES_PER_RACK) for s in completed_stores), 2),
        "collected_pieces": round(sum(num(s.get("collected_pieces")) for s in completed_stores), 2),
    }

    return render_template(
        "archive.html",
        stores=completed_stores[:500],
        routes=completed_routes[:200],
        totals=totals,
        city_options=city_options,
        driver_options=driver_options,
        filters={"q": q, "city": city, "driver": driver, "date_from": date_from, "date_to": date_to, "type": archive_type},
        can_export=is_admin(),
        can_edit_closeout=current_role() in {"Admin", "Operations Manager", "Dispatcher"},
    )

@app.route("/api/store-closeout/<store_id>", methods=["POST"])
@dispatch_required
def api_store_closeout_update(store_id):
    data = request.get_json(force=True)
    stores = read_json(STORES_FILE)
    updated = None
    for store in stores:
        if store.get("id") == store_id:
            if "collected_racks" in data and data.get("collected_racks") not in (None, ""):
                collected_racks = num(data.get("collected_racks"))
                if collected_racks < 0:
                    return jsonify({"ok": False, "message": "Collected rack count cannot be negative."}), 400
                store["collected_racks"] = collected_racks
                store["variance"] = collected_racks - num(store.get("expected_racks"))
                store["variance_review"] = abs(num(store.get("variance"))) >= 2
            if "collected_pieces" in data and data.get("collected_pieces") not in (None, ""):
                collected_pieces = num(data.get("collected_pieces"))
                if collected_pieces < 0:
                    return jsonify({"ok": False, "message": "PCS count cannot be negative."}), 400
                store["collected_pieces"] = collected_pieces
                expected_pieces = num(store.get("expected_pieces")) or (num(store.get("expected_racks")) * PIECES_PER_RACK)
                store["pieces_variance"] = collected_pieces - expected_pieces
            store["closeout_updated_at"] = datetime.now().isoformat(timespec="seconds")
            store["closeout_updated_by"] = session.get("username", "system")
            updated = store
            break
    if not updated:
        return jsonify({"ok": False, "message": "BOL/store not found."}), 404
    write_json(STORES_FILE, stores)

    routes = read_json(ROUTES_FILE)
    for route in routes:
        route_store_ids = set(route.get("store_ids", []))
        if store_id in route_store_ids:
            route_stores = [s for s in stores if s.get("id") in route_store_ids]
            route["completion_summary"] = completion_summary_for_stores(route_stores)
    write_json(ROUTES_FILE, routes)
    audit("Update Store Closeout", {"store_id": store_id, "collected_racks": updated.get("collected_racks"), "collected_pieces": updated.get("collected_pieces")})
    return jsonify({"ok": True, "store": updated})

@app.route("/api/system-health")
@admin_required
def api_system_health():
    settings_data = read_json(SETTINGS_FILE)
    health = {
        "ok": True,
        "time": datetime.now().isoformat(timespec="seconds"),
        "paths": {
            "DATA_DIR": path_health(DATA_DIR),
            "UPLOAD_DIR": path_health(UPLOAD_DIR),
            "BOL_DIR": path_health(BOL_DIR),
            "Completed": path_health(BOL_DIR / "Completed"),
            "Dispatched": path_health(BOL_DIR / "Dispatched"),
        },
        "playwright": {
            "python_package_available": bool(PLAYWRIGHT_AVAILABLE),
            "browser_path_setting": os.environ.get("PLAYWRIGHT_BROWSERS_PATH", ""),
        },
        "rms": {
            "username_saved": bool(clean(settings_data.get("rms_username"))),
            "password_saved": bool(clean(settings_data.get("rms_password"))),
            "last_login": settings_data.get("last_rms_login", ""),
            "connection_status": settings_data.get("rms_connection_status", ""),
        },
    }
    health["ok"] = all(item.get("writable") for item in health["paths"].values()) and health["playwright"]["python_package_available"]
    return jsonify(health)

@app.route("/api/export/data")
@admin_required
def api_export_data():
    buffer = BytesIO()
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    files = [
        ("stores.json", STORES_FILE),
        ("routes.json", ROUTES_FILE),
        ("audit_log.json", AUDIT_FILE),
        ("sync_history.json", SYNC_HISTORY_FILE),
        ("rms_queue.json", RMS_QUEUE_FILE),
        ("users.json", USERS_FILE),
        ("settings.json", SETTINGS_FILE),
    ]
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, path in files:
            if path.exists():
                zf.write(path, f"data/{name}")
        zf.writestr("README.txt", "EOMS operational data export. Store securely; this may include usernames, routes, BOL metadata, and audit history.\n")
    buffer.seek(0)
    audit("Export Data", {"filename": f"eoms-data-export-{now}.zip"})
    return send_file(
        buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"eoms-data-export-{now}.zip",
    )


@app.route("/users")
@admin_required
def users_admin():
    data = users_payload()
    city_options = ["All", "San Antonio", "Houston", "Dallas", "Austin", "Killeen", "Waco", "Corpus Christi", "South Texas"]
    return render_template("users.html", users=data.get("users", []), city_options=city_options)

@app.route("/users/create", methods=["POST"])
@admin_required
def users_create():
    username = clean(request.form.get("username"))
    password = request.form.get("password") or ""
    role = clean(request.form.get("role")) or "Dispatcher"
    if role not in {"Admin", "Operations Manager", "Dispatcher", "Driver"}:
        role = "Dispatcher"
    assigned_cities = request.form.getlist("assigned_cities") or ["San Antonio"]
    data = users_payload()
    users = data.get("users", [])
    if not username or not password:
        return redirect("/users")
    if any(u.get("username") == username for u in users):
        return redirect("/users")
    if role not in {"Admin", "Operations Manager"} and "All" in assigned_cities:
        assigned_cities = [c for c in assigned_cities if c != "All"] or ["San Antonio"]
    users.append({"id": str(uuid4()), "username": username, "password": hash_password(password), "role": role, "assigned_cities": assigned_cities, "active": True, "created_at": datetime.now().isoformat(timespec="seconds")})
    data["users"] = users
    save_users_payload(data)
    audit("Create User", {"username": username, "role": role, "assigned_cities": assigned_cities})
    return redirect("/users")

@app.route("/users/update/<user_id>", methods=["POST"])
@admin_required
def users_update(user_id):
    data = users_payload()
    for user in data.get("users", []):
        if user.get("id") == user_id or user.get("username") == user_id:
            role = clean(request.form.get("role")) or user.get("role", "Dispatcher")
            user["role"] = role if role in {"Admin", "Operations Manager", "Dispatcher", "Driver"} else "Dispatcher"
            user["assigned_cities"] = request.form.getlist("assigned_cities") or user.get("assigned_cities", ["San Antonio"])
            if user["role"] not in {"Admin", "Operations Manager"} and "All" in user["assigned_cities"]:
                user["assigned_cities"] = [c for c in user["assigned_cities"] if c != "All"] or ["San Antonio"]
            user["active"] = bool(request.form.get("active"))
            new_password = request.form.get("password") or ""
            if new_password:
                user["password"] = hash_password(new_password)
            user["updated_at"] = datetime.now().isoformat(timespec="seconds")
            break
    save_users_payload(data)
    return redirect("/users")

@app.route("/users/delete/<user_id>", methods=["POST"])
@admin_required
def users_delete(user_id):
    data = users_payload()
    users = data.get("users", [])
    current = current_user() or {}
    target = next((u for u in users if u.get("id") == user_id or u.get("username") == user_id), None)

    if not target:
        audit("Delete User Failed", {"user_id": user_id, "reason": "not found"})
        return redirect("/users")

    # Prevent the current administrator from deleting their own active session.
    if target.get("id") == current.get("id") or target.get("username") == current.get("username"):
        audit("Delete User Blocked", {"username": target.get("username"), "reason": "self delete"})
        return redirect("/users")

    # Prevent deleting the last active admin account.
    active_admins = [
        u for u in users
        if u.get("role") == "Admin" and u.get("active", True)
        and not (u.get("id") == target.get("id") or u.get("username") == target.get("username"))
    ]
    if target.get("role") == "Admin" and target.get("active", True) and not active_admins:
        audit("Delete User Blocked", {"username": target.get("username"), "reason": "last active admin"})
        return redirect("/users")

    data["users"] = [
        u for u in users
        if not (u.get("id") == target.get("id") or u.get("username") == target.get("username"))
    ]
    save_users_payload(data)
    audit("Delete User", {"username": target.get("username"), "role": target.get("role")})
    return redirect("/users")

@app.route("/settings", methods=["GET", "POST"])
@admin_required
def settings():
    settings_data = read_json(SETTINGS_FILE)
    message = ""
    if request.method == "POST":
        settings_data["rms_username"] = clean(request.form.get("rms_username"))
        settings_data["rms_login_url"] = clean(request.form.get("rms_login_url")) or "https://rms.reusability.com/login"
        settings_data["rms_bol_url"] = clean(request.form.get("rms_bol_url")) or "https://rms.reusability.com/bills-of-lading"
        settings_data["azure_maps_key"] = clean(request.form.get("azure_maps_key")) or settings_data.get("azure_maps_key", "")
        map_default_type = clean(request.form.get("map_default_type")).lower()
        if map_default_type not in {"satellite", "roadmap", "hybrid", "terrain"}:
            map_default_type = "satellite"
        settings_data["map_default_type"] = map_default_type
        settings_data["map_default_zoom"] = max(3, min(18, int(num(request.form.get("map_default_zoom")) or 7)))
        settings_data["map_board_default_open"] = bool(request.form.get("map_board_default_open"))
        settings_data["map_live_refresh_seconds"] = max(10, min(300, int(num(request.form.get("map_live_refresh_seconds")) or 30)))
        settings_data["due_red_days"] = max(0, min(30, int(num(request.form.get("due_red_days")) or 4)))
        settings_data["due_amber_days"] = max(settings_data["due_red_days"], min(60, int(num(request.form.get("due_amber_days")) or 7)))
        settings_data["telnyx_from_number"] = clean(request.form.get("telnyx_from_number")) or settings_data.get("telnyx_from_number", "")
        telnyx_api_key = clean(request.form.get("telnyx_api_key"))
        if telnyx_api_key:
            settings_data["telnyx_api_key"] = telnyx_api_key
        settings_data["telnyx_api_key_saved"] = bool(settings_data.get("telnyx_api_key") or os.environ.get("TELNYX_API_KEY"))
        settings_data["remember_rms_credentials"] = bool(request.form.get("remember_rms_credentials"))

        rms_password = clean(request.form.get("rms_password"))

        if settings_data["remember_rms_credentials"]:
            if rms_password:
                settings_data["rms_password"] = rms_password
                settings_data["rms_password_saved"] = True
            elif settings_data.get("rms_password"):
                settings_data["rms_password_saved"] = True
            else:
                settings_data["rms_password_saved"] = False
        else:
            settings_data["rms_password"] = ""
            settings_data["rms_password_saved"] = False
        write_json(SETTINGS_FILE, settings_data)
        audit("Update RMS Settings", {"username": settings_data["rms_username"], "password_saved": settings_data["rms_password_saved"], "azure_maps_key_saved": bool(settings_data.get("azure_maps_key"))})
        message = "RMS, mapping, and Telnyx settings saved."
    return render_template("settings.html", settings=settings_data, message=message)



@app.route("/api/dispatch-board-live")
def api_dispatch_board_live():
    stores = filter_stores_for_user(read_json(STORES_FILE))
    active_statuses = {"Need Review", "Unassigned", "Assigned", "Dispatched"}
    active = [s for s in stores if s.get("status", "Unassigned") in active_statuses]
    completed_today = [
        s for s in stores
        if s.get("status") == "Completed" and date_value(s.get("completed_at")) == today_iso()
    ]
    board_stores = active + completed_today
    racks = round(sum(num(s.get("expected_racks")) for s in active), 1)
    weight = round(sum(num(s.get("weight")) for s in active), 1)
    pieces = round(racks * PIECES_PER_RACK, 1)
    return jsonify({
        "ok": True,
        "stores": board_stores,
        "metrics": {
            "stores": len(active),
            "racks": racks,
            "pieces": pieces,
            "weight": weight,
            "revenue": round(pieces * RATE_PER_PIECE, 2),
            "driver_pay": round(pieces * DRIVER_PAY_PER_PIECE, 2),
            "max_payload": MAX_PAYLOAD,
            "remaining_capacity": round(MAX_PAYLOAD - weight, 1),
        }
    })


@app.route("/api/send-route-sms", methods=["POST"])
@dispatch_required
def api_send_route_sms():
    data = request.get_json(force=True)
    route_id = data.get("route_id")
    routes = read_json(ROUTES_FILE)
    route = next((r for r in routes if r.get("id") == route_id or r.get("route_number") == route_id), None)
    if not route:
        return jsonify({"ok": False, "message": "Route not found."}), 404

    phone = clean(route.get("driver_phone"))
    if not phone and clean(route.get("driver")):
        driver_name = clean(route.get("driver"))
        for user in users_payload().get("users", []):
            names = {clean(user.get("username")), clean(user.get("display_name"))}
            if (user.get("role") or "").lower() == "driver" and driver_name in names:
                phone = clean(user.get("phone"))
                if phone:
                    route["driver_phone"] = phone
                break
    if not phone:
        return jsonify({"ok": False, "message": "Add the driver phone number first."})

    public_base = clean(os.environ.get("PUBLIC_BASE_URL")) or request.host_url.rstrip("/")
    route_link = f"{public_base}/route-view/{route.get('id')}"
    body = f"EOMS Route {route.get('route_number')}"
    if route.get("driver"):
        body += f" for {route.get('driver')}"
    body += "\n"
    if route.get("truck"):
        body += f"Truck: {route.get('truck')}\n"
    if route.get("helper"):
        body += f"Helper: {route.get('helper')}\n"
    body += f"Stops: {len(route.get('stops', []))} | Racks: {route.get('metrics', {}).get('racks', 0)} | Weight: {route.get('metrics', {}).get('weight', 0)} lbs\nOpen route: {route_link}"

    telnyx_settings = read_json(SETTINGS_FILE)
    api_key = clean(os.environ.get("TELNYX_API_KEY")) or clean(telnyx_settings.get("telnyx_api_key"))
    from_number = clean(os.environ.get("TELNYX_FROM_NUMBER")) or clean(telnyx_settings.get("telnyx_from_number"))
    if not api_key or not from_number:
        return jsonify({"ok": False, "message": "Telnyx is not configured locally. Copy Message or set TELNYX_API_KEY and TELNYX_FROM_NUMBER.", "body": body})

    try:
        resp = requests.post(
            "https://api.telnyx.com/v2/messages",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"from": from_number, "to": phone, "text": body},
            timeout=15,
        )
        if resp.status_code >= 300:
            return jsonify({"ok": False, "message": f"Telnyx error {resp.status_code}: {resp.text[:180]}", "body": body})
        route["last_sms_sent_at"] = datetime.now().isoformat(timespec="seconds")
        write_json(ROUTES_FILE, routes)
        audit("Send Route SMS", {"route_id": route.get("id"), "route_number": route.get("route_number"), "to": phone})
        return jsonify({"ok": True, "message": "Route SMS sent.", "body": body})
    except Exception as exc:
        return jsonify({"ok": False, "message": "SMS send failed: " + str(exc)[:180], "body": body})

@app.route("/api/routes")
def api_routes():
    routes = read_json(ROUTES_FILE)
    return jsonify({"ok": True, "routes": routes})

@app.route("/api/route/<route_id>")
def api_route_detail(route_id):
    routes = read_json(ROUTES_FILE)
    for route in routes:
        if route.get("id") == route_id or route.get("route_number") == route_id:
            if current_role() == "Driver":
                user = current_user() or {}
                driver_names = {clean(user.get("username")), clean(user.get("display_name"))}
                if clean(route.get("driver")) not in driver_names:
                    return jsonify({"ok": False, "message": "Route not assigned to you."}), 403
            return jsonify({"ok": True, "route": route})
    return jsonify({"ok": False, "message": "Route not found."}), 404

@app.route("/api/preview-route", methods=["POST"])
@dispatch_required
def api_preview_route():
    data = request.get_json(force=True)
    store_ids = data.get("store_ids", [])
    mode = data.get("mode", "optimized")

    allowed_ids = {s.get("id") for s in filter_stores_for_user(read_json(STORES_FILE))}
    if any(store_id not in allowed_ids for store_id in store_ids):
        return jsonify({"ok": False, "message": "One or more stores are outside your assigned cities."}), 403

    hub_name, ordered, metrics = build_route_order(store_ids, mode)

    if not ordered:
        return jsonify({"ok": False, "message": "No unassigned stores selected."})

    return jsonify({
        "ok": True,
        "hub": hub_name,
        "mode": mode,
        "stores": ordered,
        "metrics": metrics
    })

@app.route("/api/assign-route", methods=["POST"])
@dispatch_required
def api_assign_route():
    data = request.get_json(force=True)
    driver = clean(data.get("driver"))
    driver_phone = clean(data.get("driver_phone"))
    store_ids = data.get("store_ids", [])
    mode = data.get("mode", "optimized")

    allowed_ids = {s.get("id") for s in filter_stores_for_user(read_json(STORES_FILE))}
    if any(store_id not in allowed_ids for store_id in store_ids):
        return jsonify({"ok": False, "message": "One or more stores are outside your assigned cities."}), 403

    hub_name, ordered, metrics = build_route_order(store_ids, mode)

    if not ordered:
        return jsonify({"ok": False, "message": "No unassigned stores selected."})

    if metrics["status"] == "OVER LIMIT":
        return jsonify({"ok": False, "message": "Route is over 25,001 lbs. Remove stores before assigning."})

    stores = read_json(STORES_FILE)
    assigned_ids = [s["id"] for s in ordered]

    for store in stores:
        if store["id"] in assigned_ids:
            store["status"] = "Assigned"
            store["assigned_driver"] = driver
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Assigned")

    routes = read_json(ROUTES_FILE)
    route_number = f"RT-{len(routes) + 1:05d}"

    route = {
        "id": str(uuid4()),
        "route_number": route_number,
        "driver": driver,
        "driver_phone": driver_phone,
        "hub": hub_name,
        "mode": mode,
        "mode_label": "Selection Order" if mode == "selection" else "Optimized Nearest Stop",
        "store_ids": assigned_ids,
        "stops": ordered,
        "metrics": metrics,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "status": "Assigned"
    }

    routes.append(route)
    write_json(STORES_FILE, stores)
    write_json(ROUTES_FILE, routes)
    audit("Assign Route", {"route_number": route_number, "driver": driver, "stores": len(ordered), "mode": mode})

    return jsonify({"ok": True, "route": route, "assigned": ordered, "metrics": metrics})


@app.route("/api/dispatch-route", methods=["POST"])
@dispatch_required
def api_dispatch_route():
    data = request.get_json(force=True)
    route_id = clean(data.get("route_id"))
    routes = read_json(ROUTES_FILE)
    visible_route_ids = {
        clean(r.get("id")) for r in filter_routes_for_user(routes)
    } | {
        clean(r.get("route_number")) for r in filter_routes_for_user(routes)
    }
    if route_id not in visible_route_ids:
        return jsonify({"ok": False, "message": "Route not found or outside your assigned cities."}), 404

    route = next((r for r in routes if clean(r.get("id")) == route_id or clean(r.get("route_number")) == route_id), None)
    if not route:
        return jsonify({"ok": False, "message": "Route not found."}), 404
    if not clean(route.get("driver")):
        return jsonify({"ok": False, "message": "Assign a driver before dispatch."})
    if num((route.get("metrics") or {}).get("weight")) > MAX_PAYLOAD:
        return jsonify({"ok": False, "message": "Route exceeds the 25,001 lb payload limit."})

    route["status"] = "Dispatched"
    route["dispatched_at"] = datetime.now().isoformat(timespec="seconds")
    route_store_ids = set(route.get("store_ids", []))
    stores = read_json(STORES_FILE)
    for store in stores:
        if store.get("id") in route_store_ids:
            store["status"] = "Dispatched"
            store["dispatched_at"] = route["dispatched_at"]
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Dispatched")

    write_json(ROUTES_FILE, routes)
    write_json(STORES_FILE, stores)
    audit("Dispatch Route", {"route_id": route.get("id"), "route_number": route.get("route_number"), "driver": route.get("driver")})
    return jsonify({"ok": True, "route": route, "message": "Route dispatched."})




@app.route("/api/update-route-driver", methods=["POST"])
def api_update_route_driver():
    data = request.get_json(force=True)
    route_id = data.get("route_id")
    driver = clean(data.get("driver"))
    driver_phone = clean(data.get("driver_phone"))
    truck = clean(data.get("truck"))
    helper = clean(data.get("helper"))

    routes = read_json(ROUTES_FILE)
    stores = read_json(STORES_FILE)

    updated = None
    for route in routes:
        if route.get("id") == route_id or route.get("route_number") == route_id:
            route["driver"] = driver
            route["driver_phone"] = driver_phone
            route["truck"] = truck
            route["helper"] = helper
            route["updated_at"] = datetime.now().isoformat(timespec="seconds")
            updated = route

            route_store_ids = set(route.get("store_ids", []))
            for store in stores:
                if store.get("id") in route_store_ids:
                    store["assigned_driver"] = driver
                    store["driver_phone"] = driver_phone
                    store["truck"] = truck
                    store["helper"] = helper
                    store["status"] = "Assigned"
                    if store.get("pdf_path"):
                        store["pdf_path"] = move_pdf(store["pdf_path"], "Assigned")
            break

    if not updated:
        return jsonify({"ok": False, "message": "Route not found."})

    write_json(ROUTES_FILE, routes)
    write_json(STORES_FILE, stores)
    audit("Update Route Driver", {"route_id": route_id, "driver": driver, "driver_phone": driver_phone, "truck": truck, "helper": helper})

    return jsonify({"ok": True, "route": updated})

@app.route("/api/complete-route", methods=["POST"])
def api_complete_route():
    data = request.get_json(force=True)
    route_id = data.get("route_id")
    force = bool(data.get("force"))

    routes = read_json(ROUTES_FILE)
    stores = read_json(STORES_FILE)

    route = None
    for r in routes:
        if r.get("id") == route_id or r.get("route_number") == route_id:
            r["status"] = "Completed"
            r["completed_at"] = datetime.now().isoformat(timespec="seconds")
            route = r
            break

    if not route:
        return jsonify({"ok": False, "message": "Route not found."})

    route_store_ids = set(route.get("store_ids", []))
    route_stores = [s for s in stores if s.get("id") in route_store_ids]
    missing_closeout = [
        s for s in route_stores
        if s.get("collected_racks") in (None, "")
    ]

    if missing_closeout and not force:
        return jsonify({
            "ok": False,
            "code": "CLOSEOUT_REQUIRED",
            "message": f"{len(missing_closeout)} stop(s) do not have collected rack counts. Complete driver closeout first or confirm manager override.",
            "missing": [{"bol": s.get("bol"), "store": s.get("store_name"), "city": s.get("city")} for s in missing_closeout[:10]],
        })

    for store in stores:
        if store.get("id") in route_store_ids:
            store["status"] = "Completed"
            store["completed_at"] = route.get("completed_at")
            if clean(store.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}:
                store["closed_source"] = "Both"
            else:
                store["closed_source"] = "EOMS"
            if store.get("collected_racks") in (None, ""):
                store["collected_racks"] = num(store.get("expected_racks"))
                store["variance"] = 0
                store["closeout_override"] = True
            if store.get("collected_pieces") in (None, ""):
                store["collected_pieces"] = num(store.get("collected_racks")) * PIECES_PER_RACK
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Completed")

    route["completion_summary"] = completion_summary_for_stores([s for s in stores if s.get("id") in route_store_ids])

    write_json(ROUTES_FILE, routes)
    write_json(STORES_FILE, stores)
    audit("Complete Route", {"route_id": route_id, "route_number": route.get("route_number")})

    return jsonify({"ok": True, "route": route})

@app.route("/route-view/<route_id>")
def route_view(route_id):
    routes = read_json(ROUTES_FILE)
    route = None
    for r in routes:
        if r.get("id") == route_id or r.get("route_number") == route_id:
            route = r
            break

    if not route:
        return "Route not found", 404

    if current_role() == "Driver":
        user = current_user() or {}
        driver_names = {clean(user.get("username")), clean(user.get("display_name"))}
        if clean(route.get("driver")) not in driver_names:
            return render_template("access_denied.html"), 403

    return render_template("route_view.html", route=route)

@app.route("/route-print/<route_id>")
def route_print(route_id):
    routes = read_json(ROUTES_FILE)
    route = None
    for r in routes:
        if r.get("id") == route_id or r.get("route_number") == route_id:
            route = r
            break

    if not route:
        return "Route not found", 404

    if current_role() == "Driver":
        user = current_user() or {}
        driver_names = {clean(user.get("username")), clean(user.get("display_name"))}
        if clean(route.get("driver")) not in driver_names:
            return render_template("access_denied.html"), 403

    return render_template("route_print.html", route=route)

@app.route("/api/unassign-route", methods=["POST"])
def api_unassign_route():
    data = request.get_json(force=True)
    route_id = data.get("route_id")

    routes = read_json(ROUTES_FILE)
    stores = read_json(STORES_FILE)

    target_route = None
    remaining_routes = []

    for route in routes:
        if route.get("id") == route_id or route.get("route_number") == route_id:
            target_route = route
        else:
            remaining_routes.append(route)

    if not target_route:
        return jsonify({"ok": False, "message": "Route not found."})

    route_store_ids = set(target_route.get("store_ids", []))

    restored = 0
    for store in stores:
        if store.get("id") in route_store_ids:
            store["status"] = "Unassigned"
            store["assigned_driver"] = ""
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Imported")
            restored += 1

    write_json(STORES_FILE, stores)
    write_json(ROUTES_FILE, remaining_routes)
    audit("Unassign Entire Route", {"route_id": route_id, "restored": restored})

    return jsonify({"ok": True, "message": f"Route unassigned. {restored} stores returned to Dispatch Map."})

@app.route("/api/bol/<store_id>", methods=["POST"])
@dispatch_required
def api_update_bol(store_id):
    data = request.get_json(force=True)
    allowed_statuses = {"Need Review", "Unassigned", "Assigned", "Dispatched", "Completed"}
    editable = [
        "bol", "origin", "store_name", "address", "city", "state", "zip",
        "contact", "hub", "due_date", "assigned_date", "expected_racks",
        "expected_pieces", "corner_posts", "drb40", "drb48", "wood_shelf",
        "weight", "status", "dispatch_group"
    ]

    stores = read_json(STORES_FILE)
    updated = None
    for store in stores:
        if store.get("id") == store_id or clean(store.get("bol")) == store_id:
            before_location = tuple(clean(store.get(k)) for k in ("address", "city", "state", "zip", "hub"))
            for key in editable:
                if key not in data:
                    continue
                value = clean(data.get(key))
                if key in {"expected_racks", "expected_pieces", "corner_posts", "drb40", "drb48", "wood_shelf", "weight"}:
                    store[key] = num(value)
                elif key == "status":
                    if value in allowed_statuses:
                        store[key] = value
                else:
                    store[key] = value
            apply_material_counts(store, data)

            after_location = tuple(clean(store.get(k)) for k in ("address", "city", "state", "zip", "hub"))
            if after_location != before_location:
                lat, lng, geocode_status, full_address = resolve_store_coordinates(
                    store.get("address"), store.get("city"), store.get("state") or "TX", store.get("zip")
                )
                store["lat"] = lat
                store["lng"] = lng
                store["full_address"] = full_address
                store["geocode_status"] = geocode_status
                if not clean(data.get("hub")):
                    hub, hub_reason = assign_hub(lat, lng, store.get("dispatch_group"))
                    store["hub"] = hub
                    store["hub_reason"] = hub_reason

            if clean(store.get("status")) != "Completed":
                store["review_reasons"] = essential_review_reasons(store)
                if store["review_reasons"] and clean(store.get("status")) == "Unassigned":
                    store["status"] = "Need Review"
            store["updated_at"] = datetime.now().isoformat(timespec="seconds")
            updated = store
            break

    if not updated:
        queue = read_json(RMS_QUEUE_FILE)
        for q in queue:
            if q.get("id") == store_id or clean(q.get("bol")) == store_id:
                for key in editable:
                    if key not in data:
                        continue
                    value = clean(data.get(key))
                    if key in {"expected_racks", "expected_pieces", "corner_posts", "drb40", "drb48", "wood_shelf", "weight"}:
                        q[key] = num(value)
                    elif key == "status":
                        q["queue_status"] = value or q.get("queue_status") or "New"
                    else:
                        q[key] = value
                apply_material_counts(q, data)
                q["updated_at"] = datetime.now().isoformat(timespec="seconds")
                updated = q
                break

        if not updated:
            return jsonify({"ok": False, "message": "BOL not found."}), 404

        write_json(RMS_QUEUE_FILE, queue)
        audit("Update Queue BOL", {"queue_id": updated.get("id"), "bol": updated.get("bol"), "status": updated.get("queue_status")})
        updated["status"] = updated.get("queue_status", updated.get("status", "New"))
        return jsonify({"ok": True, "store": updated})

    write_json(STORES_FILE, stores)

    queue = read_json(RMS_QUEUE_FILE)
    for q in queue:
        if clean(q.get("bol")) == clean(updated.get("bol")):
            for key in editable:
                if key in updated:
                    q[key] = updated.get(key)
            q["queue_status"] = "Imported" if updated.get("status") != "Need Review" else "Need Review"
            q["updated_at"] = updated.get("updated_at")
    write_json(RMS_QUEUE_FILE, queue)

    audit("Update BOL", {"store_id": updated.get("id"), "bol": updated.get("bol"), "status": updated.get("status")})
    sync_route_stop_materials(updated)
    return jsonify({"ok": True, "store": updated})


@app.route("/api/store-status", methods=["POST"])
def api_store_status():
    data = request.get_json(force=True)
    store_id = data.get("store_id")
    new_status = clean(data.get("status"))

    allowed = {"Need Review", "Unassigned", "Assigned", "Dispatched", "Completed"}
    if new_status not in allowed:
        return jsonify({"ok": False, "message": "Invalid status."})

    stores = read_json(STORES_FILE)
    updated = None

    for store in stores:
        if store.get("id") == store_id:
            store["status"] = new_status
            store["updated_at"] = datetime.now().isoformat(timespec="seconds")

            if new_status == "Unassigned":
                store["assigned_driver"] = ""
                if store.get("pdf_path"):
                    store["pdf_path"] = move_pdf(store["pdf_path"], "Imported")
            elif new_status == "Completed":
                if clean(store.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}:
                    store["closed_source"] = "Both"
                else:
                    store["closed_source"] = "EOMS"
                if store.get("pdf_path"):
                    store["pdf_path"] = move_pdf(store["pdf_path"], "Completed")
            elif new_status == "Assigned":
                if store.get("pdf_path"):
                    store["pdf_path"] = move_pdf(store["pdf_path"], "Assigned")

            updated = store
            break

    if not updated:
        return jsonify({"ok": False, "message": "Store not found."})

    write_json(STORES_FILE, stores)
    audit("Update Store Status", {"store_id": store_id, "status": new_status})

    return jsonify({"ok": True, "store": updated})

@app.route("/api/unassign-store", methods=["POST"])
def api_unassign_store():
    data = request.get_json(force=True)
    store_id = data.get("store_id")
    stores = read_json(STORES_FILE)
    restored = None
    for store in stores:
        if store.get("id") == store_id:
            store["status"] = "Unassigned"
            store["assigned_driver"] = ""
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Imported")
            restored = store
            break

    routes = read_json(ROUTES_FILE)
    cleaned_routes = []
    for route in routes:
        route["store_ids"] = [sid for sid in route.get("store_ids", []) if sid != store_id]
        route["stops"] = [s for s in route.get("stops", []) if s.get("id") != store_id]
        if route.get("stops"):
            route["metrics"] = calculate_route_metrics(route["stops"], route.get("hub") or "San Antonio")
            cleaned_routes.append(route)

    write_json(STORES_FILE, stores)
    write_json(ROUTES_FILE, cleaned_routes)
    audit("Unassign Store", {"store_id": store_id})
    return jsonify({"ok": True, "store": restored})

@app.route("/driver")
def driver_portal():
    stores = [s for s in read_json(STORES_FILE) if s.get("status") in {"Assigned", "Dispatched"}]
    if current_role() == "Driver":
        user = current_user() or {}
        driver_names = {clean(user.get("username")), clean(user.get("display_name"))}
        stores = [s for s in stores if clean(s.get("assigned_driver")) in driver_names]
    return render_template("driver.html", stores=stores)

@app.route("/api/driver/complete", methods=["POST"])
def api_driver_complete():
    data = request.get_json(force=True)
    store_id = data.get("store_id")
    if data.get("collected_racks") in (None, ""):
        return jsonify({"ok": False, "message": "Enter collected rack count before completing this stop."}), 400
    collected_racks = num(data.get("collected_racks"))
    if collected_racks < 0:
        return jsonify({"ok": False, "message": "Collected rack count cannot be negative."}), 400
    collected_pieces = None
    if data.get("collected_pieces") not in (None, ""):
        collected_pieces = num(data.get("collected_pieces"))
        if collected_pieces < 0:
            return jsonify({"ok": False, "message": "PCS count cannot be negative."}), 400
    stores = read_json(STORES_FILE)
    updated = None
    for store in stores:
        if store.get("id") == store_id:
            if current_role() == "Driver":
                user = current_user() or {}
                driver_names = {clean(user.get("username")), clean(user.get("display_name"))}
                if clean(store.get("assigned_driver")) not in driver_names:
                    return jsonify({"ok": False, "message": "This stop is not assigned to you."}), 403
            apply_material_counts(store, data)
            store["collected_racks"] = collected_racks
            store["collected_pieces"] = collected_pieces if collected_pieces is not None else collected_racks * PIECES_PER_RACK
            store["variance"] = collected_racks - num(store.get("expected_racks"))
            expected_pieces = num(store.get("expected_pieces")) or (num(store.get("expected_racks")) * PIECES_PER_RACK)
            store["pieces_variance"] = num(store.get("collected_pieces")) - expected_pieces
            store["status"] = "Completed"
            store["completed_at"] = datetime.now().isoformat(timespec="seconds")
            if clean(store.get("rms_status")) in {"Missing from RMS", "Closed in RMS"}:
                store["closed_source"] = "Both"
            else:
                store["closed_source"] = "EOMS"
            if abs(num(store.get("variance"))) >= 2:
                store["variance_review"] = True
            if store.get("pdf_path"):
                store["pdf_path"] = move_pdf(store["pdf_path"], "Completed")
            updated = store
            break
    write_json(STORES_FILE, stores)
    routes = read_json(ROUTES_FILE)
    for route in routes:
        route_store_ids = set(route.get("store_ids", []))
        if store_id in route_store_ids:
            route_stores = [s for s in stores if s.get("id") in route_store_ids]
            if route_stores and all((s.get("status") == "Completed") for s in route_stores):
                route["status"] = "Completed"
                route["completed_at"] = datetime.now().isoformat(timespec="seconds")
                route["completion_summary"] = completion_summary_for_stores(route_stores)
    write_json(ROUTES_FILE, routes)
    if updated:
        sync_route_stop_materials(updated)
    audit("Driver Complete", {"store_id": store_id, "collected_racks": collected_racks, "collected_pieces": collected_pieces})
    if not updated:
        return jsonify({"ok": False, "message": "Stop not found."}), 404
    return jsonify({"ok": True, "store": updated})

# Run startup tasks at import time so this works under gunicorn (which imports
# `app:app` and never executes the __main__ block below).
ensure_dirs()
migrate_user_passwords()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
